"""bank_rec.py -- Bank reconciliation engine for MEC Automation Tool.

Loads a bank statement and GL cash subledger, matches transactions by
amount and date, then categorizes unmatched items as deposits in transit,
outstanding checks, or unadjusted book/bank entries.

Public API:
    load_bank_statement(path)             -> DataFrame
    load_gl_subledger(path)               -> DataFrame
    reconcile(bank_df, gl_df,
              bank_ending_balance,
              gl_ending_balance,
              date_tolerance_days=3)      -> dict
    export_reconciliation(result, path)   -> Path

Supported file formats: .xlsx, .xls, .csv
Column names are normalized automatically -- no template required.
"""

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Column aliases (normalized: lowercase, spaces -> underscores)
# ---------------------------------------------------------------------------

_DATE_ALIASES = [
    "date", "transaction_date", "trans_date", "posting_date", "post_date",
    "value_date", "effective_date", "settlement_date", "txn_date", "trx_date",
    "entry_date",
]
_DESC_ALIASES = [
    "description", "desc", "memo", "narrative", "details",
    "transaction_description", "payee", "reference", "transaction_details",
    "particulars", "remarks", "note", "notes",
]
_AMOUNT_ALIASES = [
    "amount", "net_amount", "transaction_amount", "txn_amount", "trx_amount",
    "net",
]
_DEBIT_ALIASES = [
    "debit", "withdrawals", "withdrawal", "payment", "payments", "charges",
    "debit_amount", "withdrawal_amount", "dr",
]
_CREDIT_ALIASES = [
    "credit", "deposits", "deposit", "credits", "credit_amount",
    "deposit_amount", "cr",
]
_BALANCE_ALIASES = [
    "balance", "running_balance", "ledger_balance", "available_balance",
    "ending_balance", "closing_balance",
]
_GL_REF_ALIASES = [
    "reference", "ref", "check_number", "check_no", "doc_number",
    "transaction_ref", "voucher", "check_num",
]

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_NAVY    = "1F497D"
_RED_BG  = "FFC7CE"
_GRN_BG  = "E2EFDA"
_YEL_BG  = "FFEB9C"
_BLU_LT  = "EBF3FB"
_GRY_BG  = "F2F2F2"

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _norm(s: str) -> str:
    return re.sub(r"[\s\-./]+", "_", str(s).strip().lower())


def _find_col(df: pd.DataFrame, aliases: list) -> str | None:
    normed = {_norm(c): c for c in df.columns}
    for alias in aliases:
        if alias in normed:
            return normed[alias]
    return None


def _to_float(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.astype(str)
              .str.replace(",", "", regex=False)
              .str.replace("$", "", regex=False)
              .str.replace("(", "-", regex=False)
              .str.replace(")", "", regex=False),
        errors="coerce",
    ).fillna(0.0)


def _try_load_csv(path: Path) -> pd.DataFrame:
    for sep in (",", ";", "\t", "|"):
        try:
            df = pd.read_csv(path, sep=sep, dtype=str, encoding="utf-8-sig")
            if len(df.columns) > 1:
                return df
        except Exception:
            continue
    return pd.read_csv(path, dtype=str, encoding="utf-8-sig")


def _load_raw(path: Path) -> pd.DataFrame:
    suffix = path.suffix.lower()
    if suffix in (".xlsx", ".xls", ".xlsm"):
        raw = pd.read_excel(path, dtype=str)
    elif suffix == ".csv":
        raw = _try_load_csv(path)
    else:
        raise ValueError(
            f"Unsupported file type: {suffix!r}. Use .xlsx, .xls, or .csv"
        )
    raw.columns = [str(c).strip() for c in raw.columns]
    return raw.dropna(how="all").reset_index(drop=True)


# ---------------------------------------------------------------------------
# Public loaders
# ---------------------------------------------------------------------------

def load_bank_statement(path) -> pd.DataFrame:
    """Load a bank statement from Excel or CSV.

    Normalizes to: Date (datetime64), Description (str), Amount (float).
    Convention: deposits/credits are positive, withdrawals/debits are negative.

    Supports a single Amount column OR separate Debit / Credit columns.
    """
    path = Path(path)
    raw  = _load_raw(path)

    date_col = _find_col(raw, _DATE_ALIASES)
    if date_col is None:
        raise ValueError(
            f"No date column found in {path.name}. "
            f"Expected one of: {', '.join(_DATE_ALIASES[:6])}"
        )

    out = pd.DataFrame()
    out["Date"] = pd.to_datetime(raw[date_col], errors="coerce")
    out = out.dropna(subset=["Date"])

    desc_col = _find_col(raw, _DESC_ALIASES)
    out["Description"] = raw[desc_col].fillna("").astype(str) if desc_col else ""

    amount_col = _find_col(raw, _AMOUNT_ALIASES)
    debit_col  = _find_col(raw, _DEBIT_ALIASES)
    credit_col = _find_col(raw, _CREDIT_ALIASES)

    if amount_col:
        out["Amount"] = _to_float(raw[amount_col])
    elif debit_col and credit_col:
        out["Amount"] = _to_float(raw[credit_col]) - _to_float(raw[debit_col])
    elif debit_col:
        out["Amount"] = -_to_float(raw[debit_col])
    elif credit_col:
        out["Amount"] = _to_float(raw[credit_col])
    else:
        raise ValueError(
            f"No amount column found in {path.name}. "
            "Expected 'Amount' or separate 'Debit'/'Credit' columns."
        )

    bal_col = _find_col(raw, _BALANCE_ALIASES)
    if bal_col:
        out["Balance"] = _to_float(raw[bal_col])

    out["Source"] = "Bank"
    return out.reset_index(drop=True)


def load_gl_subledger(path) -> pd.DataFrame:
    """Load a GL cash subledger from Excel or CSV.

    Normalizes to: Date (datetime64), Description (str), Amount (float),
    Reference (str).
    Convention: same as bank statement (credits positive, debits negative).
    """
    path = Path(path)
    raw  = _load_raw(path)

    date_col = _find_col(raw, _DATE_ALIASES)
    if date_col is None:
        raise ValueError(
            f"No date column found in {path.name}. "
            f"Expected one of: {', '.join(_DATE_ALIASES[:6])}"
        )

    out = pd.DataFrame()
    out["Date"] = pd.to_datetime(raw[date_col], errors="coerce")
    out = out.dropna(subset=["Date"])

    desc_col = _find_col(raw, _DESC_ALIASES)
    out["Description"] = raw[desc_col].fillna("").astype(str) if desc_col else ""

    ref_col = _find_col(raw, _GL_REF_ALIASES)
    out["Reference"] = raw[ref_col].fillna("").astype(str) if ref_col else ""

    amount_col = _find_col(raw, _AMOUNT_ALIASES)
    debit_col  = _find_col(raw, _DEBIT_ALIASES)
    credit_col = _find_col(raw, _CREDIT_ALIASES)

    if amount_col:
        out["Amount"] = _to_float(raw[amount_col])
    elif debit_col and credit_col:
        out["Amount"] = _to_float(raw[credit_col]) - _to_float(raw[debit_col])
    elif debit_col:
        out["Amount"] = -_to_float(raw[debit_col])
    elif credit_col:
        out["Amount"] = _to_float(raw[credit_col])
    else:
        raise ValueError(
            f"No amount column found in {path.name}. "
            "Expected 'Amount' or separate 'Debit'/'Credit' columns."
        )

    out["Source"] = "GL"
    return out.reset_index(drop=True)


# ---------------------------------------------------------------------------
# Reconciliation engine
# ---------------------------------------------------------------------------

def reconcile(
    bank_df: pd.DataFrame,
    gl_df: pd.DataFrame,
    bank_ending_balance: float,
    gl_ending_balance: float,
    date_tolerance_days: int = 3,
) -> dict:
    """Match bank and GL transactions; produce a full reconciliation result.

    Matching runs in two passes:
      Pass 1 -- exact amount + date within tolerance (marks as 'Exact' or 'Timing')
      Pass 2 -- amount match only, any date (marks as 'Amount Only')

    Unmatched bank items:
      Amount > 0  -->  Bank Credit (not in books)   e.g. interest income
      Amount < 0  -->  Bank Charge (not in books)   e.g. NSF fee, service charge

    Unmatched GL items:
      Amount > 0  -->  Deposit in Transit            e.g. wire received after cutoff
      Amount < 0  -->  Outstanding Check/Payment     e.g. check not yet cleared

    Reconciliation math (standard bank rec format):
      Bank ending balance
        + Deposits in transit
        + Outstanding checks  (these are negative, so effectively subtracted)
        = Adjusted bank balance

      GL ending balance
        + Bank credits not in books  (positive)
        + Bank charges not in books  (negative, effectively subtracted)
        = Adjusted book balance

      Difference = adjusted_bank_balance - adjusted_book_balance  (target: 0.00)

    Returns:
        {
          'matched':       DataFrame of paired transactions
          'bank_only':     DataFrame of unmatched bank items (with Category column)
          'gl_only':       DataFrame of unmatched GL items   (with Category column)
          'summary':       dict with all balance figures and counts
          'is_reconciled': bool  (True if |difference| < $0.01)
          'difference':    float
        }
    """
    bank = bank_df.copy().reset_index(drop=True)
    gl   = gl_df.copy().reset_index(drop=True)

    bank_used = [False] * len(bank)
    gl_used   = [False] * len(gl)
    pairs     = []

    # Pass 1: exact amount + date within tolerance
    for bi, brow in bank.iterrows():
        for gi, grow in gl.iterrows():
            if gl_used[gi]:
                continue
            if abs(brow["Amount"] - grow["Amount"]) < 0.005:
                days = abs((brow["Date"] - grow["Date"]).days)
                if days <= date_tolerance_days:
                    bank_used[bi] = True
                    gl_used[gi]   = True
                    pairs.append({
                        "Bank_Date":    brow["Date"],
                        "GL_Date":      grow["Date"],
                        "Description":  brow["Description"],
                        "GL_Reference": grow.get("Reference", ""),
                        "Amount":       brow["Amount"],
                        "Date_Diff":    days,
                        "Match_Type":   "Exact" if days == 0 else "Timing",
                    })
                    break

    # Pass 2: amount only
    for bi, brow in bank.iterrows():
        if bank_used[bi]:
            continue
        for gi, grow in gl.iterrows():
            if gl_used[gi]:
                continue
            if abs(brow["Amount"] - grow["Amount"]) < 0.005:
                bank_used[bi] = True
                gl_used[gi]   = True
                days = abs((brow["Date"] - grow["Date"]).days)
                pairs.append({
                    "Bank_Date":    brow["Date"],
                    "GL_Date":      grow["Date"],
                    "Description":  brow["Description"],
                    "GL_Reference": grow.get("Reference", ""),
                    "Amount":       brow["Amount"],
                    "Date_Diff":    days,
                    "Match_Type":   "Amount Only",
                })
                break

    matched_df = (
        pd.DataFrame(pairs)
        if pairs
        else pd.DataFrame(columns=[
            "Bank_Date", "GL_Date", "Description",
            "GL_Reference", "Amount", "Date_Diff", "Match_Type",
        ])
    )

    bank_only = bank[~pd.Series(bank_used)].copy()
    gl_only   = gl[~pd.Series(gl_used)].copy()

    bank_only["Category"] = bank_only["Amount"].apply(
        lambda x: "Bank Credit (not in books)" if x > 0 else "Bank Charge (not in books)"
    )
    gl_only["Category"] = gl_only["Amount"].apply(
        lambda x: "Deposit in Transit" if x > 0 else "Outstanding Check/Payment"
    )

    # Reconciliation math
    deposits_in_transit      = float(gl_only[gl_only["Amount"] > 0]["Amount"].sum())
    outstanding_checks       = float(gl_only[gl_only["Amount"] < 0]["Amount"].sum())
    bank_credits_not_booked  = float(bank_only[bank_only["Amount"] > 0]["Amount"].sum())
    bank_charges_not_booked  = float(bank_only[bank_only["Amount"] < 0]["Amount"].sum())

    adj_bank = bank_ending_balance + deposits_in_transit + outstanding_checks
    adj_book = gl_ending_balance   + bank_credits_not_booked + bank_charges_not_booked
    diff     = round(adj_bank - adj_book, 2)

    summary = {
        "bank_ending_balance":      bank_ending_balance,
        "deposits_in_transit":      round(deposits_in_transit,     2),
        "outstanding_checks":       round(outstanding_checks,      2),
        "adjusted_bank_balance":    round(adj_bank,                2),
        "gl_ending_balance":        gl_ending_balance,
        "bank_credits_not_booked":  round(bank_credits_not_booked, 2),
        "bank_charges_not_booked":  round(bank_charges_not_booked, 2),
        "adjusted_book_balance":    round(adj_book,                2),
        "difference":               diff,
        "total_bank_txns":          len(bank),
        "total_gl_txns":            len(gl),
        "matched_count":            len(matched_df),
        "bank_only_count":          len(bank_only),
        "gl_only_count":            len(gl_only),
    }

    return {
        "matched":       matched_df,
        "bank_only":     bank_only.reset_index(drop=True),
        "gl_only":       gl_only.reset_index(drop=True),
        "summary":       summary,
        "is_reconciled": abs(diff) < 0.01,
        "difference":    diff,
    }


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_reconciliation(
    rec_result: dict,
    output_path,
    commentary: list | None = None,
    period_label: str = "",
    client_name: str = "",
) -> Path:
    """Write a reconciliation workbook (3-4 sheets) to output_path.

    Sheets:
      1. Reconciliation Summary  -- balance math + key stats
      2. Matched Transactions    -- all matched pairs (green)
      3. Exceptions              -- unmatched bank items (red) + GL items (yellow)
      4. AI Analysis             -- (only if commentary is non-empty)
    """
    output_path = Path(output_path)
    wb = Workbook()

    _write_summary(wb.active, rec_result, period_label, client_name)
    wb.active.title = "Reconciliation Summary"

    ws2 = wb.create_sheet("Matched Transactions")
    _write_matched(ws2, rec_result["matched"])

    ws3 = wb.create_sheet("Exceptions")
    _write_exceptions(ws3, rec_result["bank_only"], rec_result["gl_only"])

    if commentary:
        ws4 = wb.create_sheet("AI Analysis")
        _write_ai_sheet(ws4, commentary)

    wb.save(str(output_path))
    return output_path


# -- Sheet writers -----------------------------------------------------------

def _hfill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _title_cell(ws, cell_ref: str, text: str, merge_to: str | None = None):
    if merge_to:
        ws.merge_cells(f"{cell_ref}:{merge_to}")
    c = ws[cell_ref]
    c.value     = text
    c.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill      = _hfill(_NAVY)
    c.alignment = Alignment(horizontal="center", vertical="center")
    return c

def _section_hdr(ws, row: int, text: str, cols: int = 4):
    ws.merge_cells(f"A{row}:{get_column_letter(cols)}{row}")
    c = ws.cell(row=row, column=1, value=text)
    c.font  = Font(name="Calibri", bold=True, size=11, color=_NAVY)
    c.fill  = _hfill(_BLU_LT)
    return c

def _fmt_money(c):
    c.number_format = '"$"#,##0.00'
    c.alignment = Alignment(horizontal="right")

def _fmt_date(c):
    c.number_format = "YYYY-MM-DD"


def _write_summary(ws, rec_result: dict, period_label: str, client_name: str):
    s      = rec_result["summary"]
    is_rec = rec_result["is_reconciled"]

    ws.row_dimensions[1].height = 26
    ws.row_dimensions[2].height = 22

    _title_cell(ws, "A1", "Bank Reconciliation", merge_to="D1")

    # Subtitle
    subtitle = f"{client_name}  |  {period_label}" if (client_name or period_label) else "Month-End Bank Rec"
    ws.merge_cells("A2:D2")
    sub = ws["A2"]
    sub.value = subtitle
    sub.font  = Font(name="Calibri", size=11, color="FFFFFF", italic=True)
    sub.fill  = _hfill(_NAVY)
    sub.alignment = Alignment(horizontal="center")

    # Status banner
    ws.merge_cells("A3:D3")
    banner = ws["A3"]
    if is_rec:
        banner.value = "RECONCILED  --  Difference: $0.00"
        banner.fill  = _hfill(_GRN_BG)
        banner.font  = Font(name="Calibri", bold=True, size=11, color="375623")
    else:
        banner.value = f"NOT RECONCILED  --  Difference: ${s['difference']:,.2f}"
        banner.fill  = _hfill(_RED_BG)
        banner.font  = Font(name="Calibri", bold=True, size=11, color="9C0006")
    banner.alignment = Alignment(horizontal="center")

    rows = [
        (None, None, False),
        ("BANK BALANCE RECONCILIATION", None, True),
        ("  Bank Statement Ending Balance",     s["bank_ending_balance"],    False),
        ("  + Deposits in Transit",             s["deposits_in_transit"],     False),
        ("  + Outstanding Checks/Payments",     s["outstanding_checks"],      False),
        ("  Adjusted Bank Balance",             s["adjusted_bank_balance"],   True),
        (None, None, False),
        ("BOOK BALANCE RECONCILIATION", None, True),
        ("  GL Cash Account Ending Balance",    s["gl_ending_balance"],       False),
        ("  + Bank Credits (not in books)",     s["bank_credits_not_booked"], False),
        ("  + Bank Charges (not in books)",     s["bank_charges_not_booked"], False),
        ("  Adjusted Book Balance",             s["adjusted_book_balance"],   True),
        (None, None, False),
        ("  DIFFERENCE (target: $0.00)",        s["difference"],              True),
        (None, None, False),
        ("STATISTICS", None, True),
        ("  Total Bank Transactions",           s["total_bank_txns"],         False),
        ("  Total GL Transactions",             s["total_gl_txns"],           False),
        ("  Matched Pairs",                     s["matched_count"],           False),
        ("  Exceptions -- Bank Only",           s["bank_only_count"],         False),
        ("  Exceptions -- GL Only",             s["gl_only_count"],           False),
    ]

    section_keys = {"BANK BALANCE RECONCILIATION", "BOOK BALANCE RECONCILIATION", "STATISTICS"}

    for r_offset, (label, value, is_sub) in enumerate(rows, start=4):
        if label is None:
            continue
        is_sec = label in section_keys
        ws.merge_cells(f"A{r_offset}:C{r_offset}")
        ca = ws.cell(row=r_offset, column=1, value=label)
        if is_sec:
            ca.font = Font(name="Calibri", bold=True, size=11, color=_NAVY)
            ca.fill = _hfill(_BLU_LT)
        elif is_sub:
            ca.font = Font(name="Calibri", bold=True)
            ca.fill = _hfill(_GRY_BG)
        else:
            ca.font = Font(name="Calibri", size=10)

        if value is not None and not is_sec:
            cb = ws.cell(row=r_offset, column=4, value=value)
            cb.font = Font(name="Calibri", bold=is_sub)
            if is_sub:
                cb.fill = _hfill(_GRY_BG)
            if isinstance(value, float):
                _fmt_money(cb)
            else:
                cb.number_format = "#,##0"
                cb.alignment = Alignment(horizontal="right")

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 16


def _write_matched(ws, df: pd.DataFrame):
    ws.merge_cells("A1:F1")
    _title_cell(ws, "A1", "Matched Transactions", merge_to="F1")
    ws.row_dimensions[1].height = 22

    if df.empty:
        ws.cell(row=2, column=1, value="No matched transactions.").font = Font(name="Calibri", italic=True, color="888888")
        return

    hdrs = ["Bank Date", "GL Date", "Description", "GL Reference", "Amount", "Match Type"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font   = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill   = _hfill(_NAVY)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="left")

    for r, row in df.iterrows():
        fill = _hfill("F5F5F5") if r % 2 else PatternFill()
        vals = [
            row.get("Bank_Date"), row.get("GL_Date"),
            str(row.get("Description", "")), str(row.get("GL_Reference", "")),
            row.get("Amount"), str(row.get("Match_Type", "")),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r + 3, column=c, value=val)
            cell.font   = Font(name="Calibri", size=10)
            cell.fill   = fill
            cell.border = _border()
            if c in (1, 2):
                _fmt_date(cell)
            if c == 5:
                _fmt_money(cell)

    for c, w in enumerate([13, 13, 42, 16, 14, 13], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _write_exceptions(ws, bank_only: pd.DataFrame, gl_only: pd.DataFrame):
    ws.merge_cells("A1:E1")
    _title_cell(ws, "A1", "Exceptions -- Items Requiring Follow-Up", merge_to="E1")
    ws.row_dimensions[1].height = 22

    row_ptr = 2
    sections = [
        ("Bank Statement Items NOT in GL Subledger", bank_only, _RED_BG),
        ("GL Items NOT on Bank Statement (Outstanding)", gl_only, _YEL_BG),
    ]

    for section_title, df, bg in sections:
        ws.merge_cells(f"A{row_ptr}:E{row_ptr}")
        sh = ws.cell(row=row_ptr, column=1, value=section_title)
        sh.font = Font(name="Calibri", bold=True, size=11, color=_NAVY)
        sh.fill = _hfill(_BLU_LT)
        row_ptr += 1

        if df.empty:
            c = ws.cell(row=row_ptr, column=1, value="None -- all items matched.")
            c.font = Font(name="Calibri", italic=True, color="888888")
            row_ptr += 2
            continue

        for col, h in enumerate(["Date", "Description", "Reference", "Amount", "Category"], 1):
            cell = ws.cell(row=row_ptr, column=col, value=h)
            cell.font   = Font(name="Calibri", bold=True, color="FFFFFF")
            cell.fill   = _hfill(_NAVY)
            cell.border = _border()
        row_ptr += 1

        for _, r in df.iterrows():
            vals = [
                r["Date"],
                str(r.get("Description", "")),
                str(r.get("Reference", "")),
                r["Amount"],
                str(r.get("Category", "")),
            ]
            for col, val in enumerate(vals, 1):
                cell = ws.cell(row=row_ptr, column=col, value=val)
                cell.font   = Font(name="Calibri", size=10)
                cell.fill   = _hfill(bg)
                cell.border = _border()
                if col == 1:
                    _fmt_date(cell)
                if col == 4:
                    _fmt_money(cell)
            row_ptr += 1

        row_ptr += 1

    for c, w in enumerate([13, 42, 14, 14, 30], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


def _write_ai_sheet(ws, commentary: list):
    ws.merge_cells("A1:E1")
    _title_cell(ws, "A1", "AI Exception Analysis", merge_to="E1")
    ws.row_dimensions[1].height = 22

    hdrs = ["Date", "Description", "Amount", "Explanation", "Suggested Action"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font   = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill   = _hfill(_NAVY)
        cell.border = _border()

    for r, item in enumerate(commentary, start=3):
        fill = _hfill("F5F5F5") if r % 2 else PatternFill()
        vals = [
            item.get("date", ""),
            item.get("description", ""),
            item.get("amount", 0.0),
            item.get("explanation", ""),
            item.get("action", ""),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font      = Font(name="Calibri", size=10)
            cell.fill      = fill
            cell.border    = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            if c == 3 and isinstance(val, float):
                _fmt_money(cell)
        ws.row_dimensions[r].height = 40

    for c, w in enumerate([13, 32, 14, 48, 30], 1):
        ws.column_dimensions[get_column_letter(c)].width = w


# ---------------------------------------------------------------------------
# Multi-account export
# ---------------------------------------------------------------------------

def export_multi_reconciliation(
    accounts: list[dict],
    output_path,
    period_label: str = "",
    client_name: str = "",
) -> Path:
    """Export a multi-account reconciliation workbook.

    Args:
        accounts: list of dicts, each with keys:
            "name"       -- account name (str)
            "result"     -- return value from reconcile()
            "commentary" -- list from generate_bank_rec_commentary() or []
        output_path: where to save the workbook
        period_label, client_name: metadata for title rows

    Sheets:
        1. All Accounts Summary  -- one row per account with status
        2-N. Per-account sheets  -- individual reconciliation detail
    """
    output_path = Path(output_path)
    wb = Workbook()

    _write_multi_summary(wb.active, accounts, period_label, client_name)
    wb.active.title = "All Accounts Summary"

    for acct in accounts:
        sheet_name = acct["name"][:31]  # Excel tab limit
        ws = wb.create_sheet(sheet_name)
        _write_summary(ws, acct["result"], period_label, acct["name"])
        if acct.get("commentary"):
            exc_ws = wb.create_sheet(f"{sheet_name[:27]} - AI")
            _write_ai_sheet(exc_ws, acct["commentary"])

    wb.save(str(output_path))
    return output_path


def _write_multi_summary(ws, accounts: list[dict], period_label: str, client_name: str):
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = "Multi-Account Bank Reconciliation"
    t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.fill  = _hfill(_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    sub = f"{client_name}  |  {period_label}" if (client_name or period_label) else ""
    ws.merge_cells("A2:H2")
    sc = ws["A2"]
    sc.value = sub
    sc.font  = Font(name="Calibri", size=11, color="FFFFFF", italic=True)
    sc.fill  = _hfill(_NAVY)
    sc.alignment = Alignment(horizontal="center")

    hdrs = ["Account", "Bank Ending", "GL Ending", "Difference", "Status",
            "Matched", "Exceptions", "AI Analysis"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=4, column=c, value=h)
        cell.font   = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill   = _hfill(_NAVY)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left")

    rec_count = 0
    for r, acct in enumerate(accounts, start=5):
        s      = acct["result"]["summary"]
        is_rec = acct["result"]["is_reconciled"]
        diff   = acct["result"]["difference"]
        bg     = _GRN_BG if is_rec else _RED_BG
        status = "Reconciled" if is_rec else "Not Reconciled"
        if is_rec:
            rec_count += 1

        vals = [
            acct["name"],
            s["bank_ending_balance"],
            s["gl_ending_balance"],
            diff,
            status,
            s["matched_count"],
            s["bank_only_count"] + s["gl_only_count"],
            "Yes" if acct.get("commentary") else "No",
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font   = Font(name="Calibri", size=10, bold=(c == 5))
            cell.fill   = _hfill(bg)
            cell.border = _border()
            if c in (2, 3, 4):
                _fmt_money(cell)
            if c in (6, 7):
                cell.alignment = Alignment(horizontal="center")
            if c == 5:
                cell.alignment = Alignment(horizontal="center")
            if c == 8:
                cell.alignment = Alignment(horizontal="center")

    # Footer summary
    total_row = 5 + len(accounts)
    ws.merge_cells(f"A{total_row}:D{total_row}")
    summary_cell = ws.cell(
        row=total_row, column=1,
        value=f"{rec_count} of {len(accounts)} accounts reconciled"
    )
    summary_cell.font = Font(name="Calibri", bold=True, size=11, color=_NAVY)
    summary_cell.fill = _hfill(_BLU_LT)
    summary_cell.border = _border()

    ws.column_dimensions["A"].width = 28
    for c, w in enumerate([16, 16, 14, 16, 10, 12, 14], 2):
        ws.column_dimensions[get_column_letter(c)].width = w


# ---------------------------------------------------------------------------
# Stand-alone test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import pandas as pd

    print("bank_rec.py -- stand-alone test\n")

    bank_rows = [
        ("2026-04-01", "Opening Balance Transfer",              50000.00),
        ("2026-04-02", "Client Payment - Invoice 2026-031",     12500.00),
        ("2026-04-03", "Client Payment - Invoice 2026-028",      8750.00),
        ("2026-04-05", "ACH Vendor XYZ Payment",               -3200.00),
        ("2026-04-07", "Client Payment - Invoice 2026-032",      5000.00),
        ("2026-04-08", "ACH Payroll Direct Deposit",           -22000.00),
        ("2026-04-10", "Wire Transfer - Equipment Purchase",   -15000.00),
        ("2026-04-12", "NSF Fee - Returned Check",                -35.00),
        ("2026-04-14", "ACH Rent Payment",                     -4500.00),
        ("2026-04-15", "Client Payment - Invoice 2026-029",      9500.00),
        ("2026-04-16", "ACH Insurance Premium",                -1850.00),
        ("2026-04-18", "Interest Income",                        127.50),
        ("2026-04-20", "Wire Receipt - Client ABC",            25000.00),
        ("2026-04-22", "ACH Utilities Payment",                  -890.00),
        ("2026-04-24", "ACH Software Subscription",              -299.00),
        ("2026-04-25", "Client Payment - Invoice 2026-033",      6000.00),
        ("2026-04-26", "ACH Office Supplies",                    -450.00),
        ("2026-04-28", "Wire Receipt - Client DEF",            18000.00),
        ("2026-04-29", "Bank Service Charge",                     -25.00),
    ]
    bank_df = pd.DataFrame(bank_rows, columns=["Date", "Description", "Amount"])
    bank_df["Date"]   = pd.to_datetime(bank_df["Date"])
    bank_df["Source"] = "Bank"

    gl_rows = [
        ("2026-04-01", "Transfer from Operating Reserve",  "DEP001",   0,       50000.00),
        ("2026-04-01", "Invoice 2026-031 - Client Pmt",    "DEP002",   0,       12500.00),
        ("2026-04-03", "Invoice 2026-028 - Client Pmt",    "DEP003",   0,        8750.00),
        ("2026-04-05", "Vendor XYZ - April Services",      "CHK1050",  3200.00,     0),
        ("2026-04-07", "Invoice 2026-032 - Client Pmt",    "DEP004",   0,        5000.00),
        ("2026-04-08", "Payroll - Bi-weekly",              "PAY001",  22000.00,     0),
        ("2026-04-10", "Equipment Purchase",               "CHK1051", 15000.00,     0),
        ("2026-04-14", "Office Lease - April",             "CHK1052",  4500.00,     0),
        ("2026-04-13", "Invoice 2026-029 - Client Pmt",    "DEP005",   0,        9500.00),
        ("2026-04-16", "Business Insurance Premium",       "ACH001",   1850.00,     0),
        ("2026-04-20", "Client ABC - Project Wire",        "WIRE001",     0,    25000.00),
        ("2026-04-22", "Electric/Gas Utilities",           "ACH002",    890.00,     0),
        ("2026-04-24", "Cloud Software - Monthly",         "ACH003",    299.00,     0),
        ("2026-04-25", "Invoice 2026-033 - Client Pmt",    "DEP006",   0,        6000.00),
        ("2026-04-26", "Office Supplies - Q2",             "CHK1053",   450.00,     0),
        ("2026-04-28", "Client DEF - Milestone Wire",      "WIRE002",      0,   18000.00),
        ("2026-04-30", "Client GHI - Final Invoice Wire",  "WIRE003",      0,   11000.00),
        ("2026-04-30", "Consultant Fee",                   "CHK1054",  3500.00,     0),
        ("2026-04-30", "IT Contractor Payment",            "CHK1055",  1200.00,     0),
    ]
    gl_df = pd.DataFrame(gl_rows, columns=["Date", "Description", "Reference", "Debit", "Credit"])
    gl_df["Date"]   = pd.to_datetime(gl_df["Date"])
    gl_df["Amount"] = gl_df["Credit"] - gl_df["Debit"]
    gl_df["Source"] = "GL"

    result = reconcile(
        bank_df, gl_df,
        bank_ending_balance=86628.50,
        gl_ending_balance=92861.00,
    )
    s = result["summary"]
    print(f"  Matched pairs:         {s['matched_count']}")
    print(f"  Bank-only exceptions:  {s['bank_only_count']}")
    print(f"  GL-only exceptions:    {s['gl_only_count']}")
    print(f"  Adjusted bank balance: ${s['adjusted_bank_balance']:,.2f}")
    print(f"  Adjusted book balance: ${s['adjusted_book_balance']:,.2f}")
    print(f"  Difference:            ${s['difference']:,.2f}")
    print(f"  Reconciled:            {result['is_reconciled']}")

    out = export_reconciliation(result, "test_bank_rec.xlsx",
                                period_label="April 2026", client_name="ABC Corporation")
    print(f"\n  Output: {out}")
