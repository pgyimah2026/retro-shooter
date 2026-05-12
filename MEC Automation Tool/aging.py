"""aging.py -- AR/AP aging analysis for MEC Automation Tool.

Accepts the summary-format aging report exported by QuickBooks, NetSuite,
Sage, or similar systems (one row per customer/vendor, one column per aging
bucket) and produces bucket totals, concentration analysis, flagged
high-risk items, and a formatted Excel workbook.

Public API:
    load_aging_report(path, report_type="AR") -> DataFrame
    analyze_aging(df, report_type="AR")       -> dict
    export_aging_excel(result, output_path,
                       commentary=None,
                       period_label="",
                       client_name="")        -> Path
"""

import re
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Column-detection helpers
# ---------------------------------------------------------------------------

def _norm(s: str) -> str:
    return re.sub(r"[\s\-./]+", "_", str(s).strip().lower())


_NAME_PATTERNS    = [r"^name$", r"customer", r"vendor", r"debtor", r"creditor",
                     r"account", r"client", r"company", r"entity", r"supplier"]
_CURRENT_PATTERNS = [r"^current", r"not_yet_due", r"^0_30", r"current_period"]
_B1_30_PATTERNS   = [r"^1_30", r"^1\s*-\s*30", r"1_to_30", r"past_due_1", r"period_1"]
_B31_60_PATTERNS  = [r"^31_60", r"31_to_60", r"past_due_31", r"period_2"]
_B61_90_PATTERNS  = [r"^61_90", r"61_to_90", r"past_due_61", r"period_3"]
_B90P_PATTERNS    = [r"91", r">_90", r"over_90", r"90_plus", r"90_and_over",
                     r"90\+", r"period_4", r"180", r"past_due_over"]
_TOTAL_PATTERNS   = [r"^total$", r"total_outstanding", r"^balance$",
                     r"amount_due", r"total_balance"]


def _find_col(df: pd.DataFrame, patterns: list[str]) -> str | None:
    normed = {_norm(c): c for c in df.columns}
    for pat in patterns:
        for key, orig in normed.items():
            if re.search(pat, key):
                return orig
    return None


def _to_float(series: pd.Series) -> pd.Series:
    return pd.to_numeric(
        series.astype(str)
              .str.replace(",", "", regex=False)
              .str.replace("$", "", regex=False)
              .str.replace("(", "-", regex=False)
              .str.replace(")", "", regex=False),
        errors="coerce",
    ).fillna(0.0)


def _try_load_csv(path: Path) -> pd.DataFrame:
    for sep in (",", ";", "\t", "|"):
        try:
            df = pd.read_csv(path, sep=sep, dtype=str, encoding="utf-8-sig")
            if len(df.columns) > 1:
                return df
        except Exception:
            continue
    return pd.read_csv(path, dtype=str, encoding="utf-8-sig")


# ---------------------------------------------------------------------------
# Public: loader
# ---------------------------------------------------------------------------

def load_aging_report(path, report_type: str = "AR") -> pd.DataFrame:
    """Load a pre-bucketed aging summary from Excel or CSV.

    Detects the name column and five aging bucket columns automatically.
    Skips total/subtotal rows (rows where the name column matches "total",
    "grand total", etc.).

    Returns a DataFrame with columns:
        Name, Current, B1_30, B31_60, B61_90, B90_Plus, Total
    """
    path   = Path(path)
    suffix = path.suffix.lower()

    if suffix in (".xlsx", ".xls", ".xlsm"):
        raw = pd.read_excel(path, dtype=str)
    elif suffix == ".csv":
        raw = _try_load_csv(path)
    else:
        raise ValueError(f"Unsupported file type: {suffix!r}. Use .xlsx or .csv")

    raw.columns = [str(c).strip() for c in raw.columns]
    raw = raw.dropna(how="all").reset_index(drop=True)

    name_col    = _find_col(raw, _NAME_PATTERNS)
    current_col = _find_col(raw, _CURRENT_PATTERNS)
    b1_30_col   = _find_col(raw, _B1_30_PATTERNS)
    b31_60_col  = _find_col(raw, _B31_60_PATTERNS)
    b61_90_col  = _find_col(raw, _B61_90_PATTERNS)
    b90p_col    = _find_col(raw, _B90P_PATTERNS)
    total_col   = _find_col(raw, _TOTAL_PATTERNS)

    if name_col is None:
        raise ValueError(
            f"Could not find name/customer/vendor column in {path.name}. "
            "Expected a column named: Name, Customer, Vendor, or similar."
        )

    # Skip total/summary rows
    skip_names = {"total", "grand total", "totals", "grand totals", "subtotal"}
    mask = ~raw[name_col].str.strip().str.lower().isin(skip_names)
    mask &= raw[name_col].str.strip().ne("")
    mask &= raw[name_col].notna()
    raw = raw[mask].copy()

    out = pd.DataFrame()
    out["Name"]     = raw[name_col].str.strip().astype(str)
    out["Current"]  = _to_float(raw[current_col]) if current_col else 0.0
    out["B1_30"]    = _to_float(raw[b1_30_col])   if b1_30_col   else 0.0
    out["B31_60"]   = _to_float(raw[b31_60_col])  if b31_60_col  else 0.0
    out["B61_90"]   = _to_float(raw[b61_90_col])  if b61_90_col  else 0.0
    out["B90_Plus"] = _to_float(raw[b90p_col])    if b90p_col    else 0.0

    if total_col:
        out["Total"] = _to_float(raw[total_col])
    else:
        out["Total"] = out[["Current", "B1_30", "B31_60", "B61_90", "B90_Plus"]].sum(axis=1)

    # Drop rows with zero total
    out = out[out["Total"].abs() > 0.005].reset_index(drop=True)

    return out


# ---------------------------------------------------------------------------
# Public: analysis
# ---------------------------------------------------------------------------

_BUCKETS = ["Current", "B1_30", "B31_60", "B61_90", "B90_Plus"]
_BUCKET_LABELS = {
    "Current":  "Current",
    "B1_30":    "1-30 Days",
    "B31_60":   "31-60 Days",
    "B61_90":   "61-90 Days",
    "B90_Plus": "90+ Days",
}


def analyze_aging(df: pd.DataFrame, report_type: str = "AR") -> dict:
    """Produce summary statistics and flagged items from an aging DataFrame.

    Returns:
        {
          "report_type":   "AR" | "AP",
          "entity_label":  "Customer" | "Vendor",
          "detail":        DataFrame (input with Overdue + Risk columns added),
          "summary": {
              "total", "by_bucket", "pct_by_bucket",
              "overdue_total", "overdue_pct",
              "entity_count", "flagged_count", "high_risk_count",
          },
          "flagged":      DataFrame (entities with any 61+ day balance),
          "top_entities": DataFrame (top 5 by Total),
          "concentration": list of {name, total, pct} for top 3,
        }
    """
    report_type  = report_type.upper()
    entity_label = "Customer" if report_type == "AR" else "Vendor"

    df = df.copy()
    grand_total  = float(df["Total"].sum())
    if grand_total == 0:
        grand_total = 1.0  # avoid division by zero

    by_bucket     = {b: float(df[b].sum()) for b in _BUCKETS}
    pct_by_bucket = {b: round(v / grand_total * 100, 1) for b, v in by_bucket.items()}

    overdue_total = float(df[["B1_30", "B31_60", "B61_90", "B90_Plus"]].sum().sum())
    overdue_pct   = round(overdue_total / grand_total * 100, 1)

    # Derived columns
    df["Overdue"]   = df[["B1_30", "B31_60", "B61_90", "B90_Plus"]].sum(axis=1)
    df["Pct_Total"] = (df["Total"] / grand_total * 100).round(1)
    df["Risk"] = "Low"
    df.loc[df["B1_30"] > 0, "Risk"] = "Watch"
    df.loc[(df["B31_60"] > 0) | (df["B61_90"] > 0), "Risk"] = "Elevated"
    df.loc[df["B90_Plus"] > 0, "Risk"] = "High"

    flagged      = df[df["Risk"].isin(["Elevated", "High"])].copy()
    flagged_count    = len(flagged)
    high_risk_count  = int((df["Risk"] == "High").sum())

    top_entities = df.nlargest(5, "Total")[["Name", "Total", "Overdue", "Pct_Total", "Risk"]].copy()

    # Concentration: top 3
    top3 = df.nlargest(3, "Total")
    concentration = [
        {
            "name":  row["Name"],
            "total": float(row["Total"]),
            "pct":   float(row["Pct_Total"]),
        }
        for _, row in top3.iterrows()
    ]

    summary = {
        "total":          round(float(df["Total"].sum()), 2),
        "by_bucket":      {b: round(v, 2) for b, v in by_bucket.items()},
        "pct_by_bucket":  pct_by_bucket,
        "overdue_total":  round(overdue_total, 2),
        "overdue_pct":    overdue_pct,
        "entity_count":   len(df),
        "flagged_count":  flagged_count,
        "high_risk_count": high_risk_count,
    }

    return {
        "report_type":   report_type,
        "entity_label":  entity_label,
        "detail":        df.reset_index(drop=True),
        "summary":       summary,
        "flagged":       flagged.reset_index(drop=True),
        "top_entities":  top_entities.reset_index(drop=True),
        "concentration": concentration,
    }


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

_NAVY   = "1F497D"
_RED_BG = "FFC7CE"
_YEL_BG = "FFEB9C"
_GRN_BG = "E2EFDA"
_BLU_LT = "EBF3FB"
_GRY    = "F2F2F2"

_RISK_COLORS = {
    "Low":      _GRN_BG,
    "Watch":    _BLU_LT,
    "Elevated": _YEL_BG,
    "High":     _RED_BG,
}


def _hfill(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)

def _border() -> Border:
    s = Side(style="thin", color="D0D0D0")
    return Border(left=s, right=s, top=s, bottom=s)

def _money(c):
    c.number_format = '"$"#,##0.00'
    c.alignment = Alignment(horizontal="right")

def _pct(c):
    c.number_format = '0.0"%"'
    c.alignment = Alignment(horizontal="right")


def export_aging_excel(
    result: dict,
    output_path,
    commentary: list | None = None,
    period_label: str = "",
    client_name: str = "",
) -> Path:
    """Write a 2-3 sheet aging workbook to output_path.

    Sheets:
      1. Aging Summary   -- bucket totals + entity detail table
      2. High Risk       -- flagged entities requiring follow-up
      3. AI Analysis     -- (only if commentary is non-empty)
    """
    output_path = Path(output_path)
    wb = Workbook()

    _write_summary_sheet(wb.active, result, period_label, client_name)
    wb.active.title = "Aging Summary"

    ws2 = wb.create_sheet("High Risk")
    _write_flagged_sheet(ws2, result)

    if commentary:
        ws3 = wb.create_sheet("AI Analysis")
        _write_ai_sheet(ws3, commentary, result["entity_label"])

    wb.save(str(output_path))
    return output_path


def _write_summary_sheet(ws, result: dict, period_label: str, client_name: str):
    s  = result["summary"]
    el = result["entity_label"]
    rt = result["report_type"]
    df = result["detail"]

    # Title
    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"{'Accounts Receivable' if rt == 'AR' else 'Accounts Payable'} Aging Report"
    t.font  = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    t.fill  = _hfill(_NAVY)
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 26

    # Subtitle
    sub = f"{client_name}  |  {period_label}" if (client_name or period_label) else ""
    ws.merge_cells("A2:H2")
    sc = ws["A2"]
    sc.value = sub
    sc.font  = Font(name="Calibri", size=11, color="FFFFFF", italic=True)
    sc.fill  = _hfill(_NAVY)
    sc.alignment = Alignment(horizontal="center")

    # Bucket summary tiles (row 4)
    labels = ["Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "TOTAL OUTSTANDING"]
    keys   = ["Current", "B1_30", "B31_60", "B61_90", "B90_Plus"]
    totals = [s["by_bucket"].get(k, 0) for k in keys] + [s["total"]]
    pcts   = [s["pct_by_bucket"].get(k, 0) for k in keys] + [100.0]
    fills  = [_GRN_BG, _BLU_LT, _YEL_BG, _RED_BG, _RED_BG, _GRY]

    for col, (lbl, val, pct, fill) in enumerate(zip(labels, totals, pcts, fills), 2):
        ws.merge_cells(f"{get_column_letter(col)}4:{get_column_letter(col)}4")
        hc = ws.cell(row=4, column=col, value=lbl)
        hc.font = Font(name="Calibri", bold=True, size=10, color=_NAVY)
        hc.fill = _hfill(fill)
        hc.alignment = Alignment(horizontal="center")
        hc.border = _border()

        vc = ws.cell(row=5, column=col, value=val)
        vc.font = Font(name="Calibri", bold=True, size=12)
        vc.fill = _hfill(fill)
        _money(vc)
        vc.border = _border()

        pc = ws.cell(row=6, column=col, value=pct / 100)
        pc.font = Font(name="Calibri", size=9, color="666666")
        pc.fill = _hfill(fill)
        pc.number_format = '0.0"%"'
        pc.alignment = Alignment(horizontal="center")
        pc.border = _border()

    for r in (4, 5, 6):
        ws.row_dimensions[r].height = 18

    # Key stats (col A, rows 4-6)
    stats = [
        (f"{el} Count:", s["entity_count"]),
        ("Flagged (61+ days):", s["flagged_count"]),
        ("High Risk (90+ days):", s["high_risk_count"]),
    ]
    for r_off, (lbl, val) in enumerate(stats):
        ws.cell(row=4 + r_off, column=1, value=lbl).font = Font(name="Calibri", bold=True, size=10)
        c = ws.cell(row=4 + r_off, column=2)
        # overwrite bucket col A with stats
        ws.cell(row=4 + r_off, column=1, value=lbl)

    ws["A4"].value = f"Total {el}s:"
    ws["A4"].font  = Font(name="Calibri", bold=True, size=10)
    ws.cell(row=4, column=1).value = f"Total {el}s:"
    ws.cell(row=5, column=1).value = f"Flagged (61+ days):"
    ws.cell(row=6, column=1).value = f"High Risk (90+ days):"
    for r, v in [(4, s["entity_count"]), (5, s["flagged_count"]), (6, s["high_risk_count"])]:
        ws.cell(row=r, column=1).font = Font(name="Calibri", bold=True, size=10, color=_NAVY)

    # Detail table header (row 8)
    hdrs = [el, "Current", "1-30 Days", "31-60 Days", "61-90 Days", "90+ Days", "Total", "Risk"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=8, column=c, value=h)
        cell.font   = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
        cell.fill   = _hfill(_NAVY)
        cell.border = _border()
        cell.alignment = Alignment(horizontal="center" if c > 1 else "left")

    # Detail rows
    bucket_cols = ["Current", "B1_30", "B31_60", "B61_90", "B90_Plus"]
    for r, (_, row) in enumerate(df.iterrows(), start=9):
        risk  = str(row.get("Risk", "Low"))
        bg    = _RISK_COLORS.get(risk, "FFFFFF")
        fill  = _hfill(bg) if risk != "Low" else (_hfill("F9F9F9") if r % 2 else PatternFill())

        ws.cell(row=r, column=1, value=str(row["Name"])).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).fill   = fill
        ws.cell(row=r, column=1).border = _border()

        for ci, bcol in enumerate(bucket_cols, 2):
            c = ws.cell(row=r, column=ci, value=float(row[bcol]))
            c.fill = fill; c.border = _border(); _money(c)
            c.font = Font(name="Calibri", size=10)

        tc = ws.cell(row=r, column=7, value=float(row["Total"]))
        tc.fill = fill; tc.border = _border(); _money(tc)
        tc.font = Font(name="Calibri", bold=True, size=10)

        rc = ws.cell(row=r, column=8, value=risk)
        rc.fill = fill; rc.border = _border()
        rc.font = Font(name="Calibri", bold=(risk in ("Elevated", "High")), size=10)
        rc.alignment = Alignment(horizontal="center")

    # Total row
    total_row = 9 + len(df)
    ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name="Calibri", bold=True)
    ws.cell(row=total_row, column=1).fill = _hfill(_GRY)
    ws.cell(row=total_row, column=1).border = _border()
    for ci, bcol in enumerate(bucket_cols, 2):
        c = ws.cell(row=total_row, column=ci, value=s["by_bucket"][bcol])
        c.font = Font(name="Calibri", bold=True); c.fill = _hfill(_GRY)
        c.border = _border(); _money(c)
    tc = ws.cell(row=total_row, column=7, value=s["total"])
    tc.font = Font(name="Calibri", bold=True); tc.fill = _hfill(_GRY)
    tc.border = _border(); _money(tc)
    ws.cell(row=total_row, column=8).fill = _hfill(_GRY)
    ws.cell(row=total_row, column=8).border = _border()

    ws.column_dimensions["A"].width = 28
    for c, w in enumerate([16, 16, 16, 16, 16, 16, 12], 2):
        ws.column_dimensions[get_column_letter(c)].width = w


def _write_flagged_sheet(ws, result: dict):
    el      = result["entity_label"]
    flagged = result["flagged"]

    ws.merge_cells("A1:H1")
    t = ws["A1"]
    t.value = f"High-Risk {el}s -- Overdue 61+ Days"
    t.font  = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    t.fill  = _hfill(_NAVY)
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    if flagged.empty:
        ws.cell(row=2, column=1, value=f"No {el.lower()}s overdue 61+ days. All accounts are current or in early aging.").font = Font(name="Calibri", italic=True, color="375623")
        return

    hdrs = [el, "61-90 Days", "90+ Days", "Total Overdue", "Total Outstanding", "Risk"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = _hfill(_NAVY)
        cell.border = _border()

    for r, (_, row) in enumerate(flagged.iterrows(), start=3):
        risk = str(row.get("Risk", "Elevated"))
        fill = _hfill(_RED_BG if risk == "High" else _YEL_BG)
        vals = [
            str(row["Name"]),
            float(row["B61_90"]),
            float(row["B90_Plus"]),
            float(row["Overdue"]),
            float(row["Total"]),
            risk,
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Calibri", size=10, bold=(c == 6 and risk == "High"))
            cell.fill = fill
            cell.border = _border()
            if c in (2, 3, 4, 5):
                _money(cell)
            if c == 6:
                cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 28
    for c, w in enumerate([16, 14, 16, 18, 10], 2):
        ws.column_dimensions[get_column_letter(c)].width = w


def _write_ai_sheet(ws, commentary: list, entity_label: str):
    ws.merge_cells("A1:D1")
    t = ws["A1"]
    t.value = f"AI Aging Analysis -- {entity_label} Collection Risk"
    t.font  = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
    t.fill  = _hfill(_NAVY)
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    hdrs = [entity_label, "Risk Level", "Analysis", "Recommended Action"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = Font(name="Calibri", bold=True, color="FFFFFF")
        cell.fill = _hfill(_NAVY)
        cell.border = _border()

    for r, item in enumerate(commentary, start=3):
        risk = item.get("risk", "")
        fill = _hfill(_RISK_COLORS.get(risk, "FFFFFF"))
        vals = [
            item.get("name", ""),
            risk,
            item.get("analysis", ""),
            item.get("action", ""),
        ]
        for c, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = Font(name="Calibri", size=10)
            cell.fill = fill
            cell.border = _border()
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[r].height = 44

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 34


# ---------------------------------------------------------------------------
# Stand-alone test
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    sample = pd.DataFrame({
        "Name":     ["Acme Industries", "BuildRight Corp", "Coastal Services",
                     "Delta Tech", "Eagle Enterprises", "First National",
                     "Granite Partners", "Harbor Solutions", "Ironclad Group",
                     "Junction LLC"],
        "Current":  [15000,  8500,     0, 22000,     0, 12000,    0,  9500, 31000,     0],
        "B1_30":    [    0,  3200,  5500,     0,     0,  6800, 4100,     0,     0,  8800],
        "B31_60":   [    0,     0,  2800,  4500,     0,     0,    0,  3600,     0,  4200],
        "B61_90":   [    0,     0,     0,     0,  7200,     0, 2900,     0,     0,  1100],
        "B90_Plus": [    0,     0,     0,     0,  3500,     0,    0,  1800,     0,  6200],
    })
    sample["Total"] = sample[["Current", "B1_30", "B31_60", "B61_90", "B90_Plus"]].sum(axis=1)

    result = analyze_aging(sample, report_type="AR")
    s = result["summary"]
    print("AR Aging Analysis")
    print(f"  Total AR:         ${s['total']:,.2f}")
    print(f"  Overdue:          ${s['overdue_total']:,.2f}  ({s['overdue_pct']:.1f}%)")
    print(f"  Flagged (61+d):   {s['flagged_count']} entities")
    print(f"  High risk (90+d): {s['high_risk_count']} entities")
    print(f"\n  Top 3 by balance:")
    for e in result["concentration"]:
        print(f"    {e['name']:<22}  ${e['total']:>10,.2f}  ({e['pct']:.1f}%)")

    out = export_aging_excel(result, "test_aging.xlsx",
                             period_label="April 2026", client_name="ABC Corporation")
    print(f"\n  Output: {out}")
