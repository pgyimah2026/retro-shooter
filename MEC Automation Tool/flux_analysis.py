"""flux_analysis.py -- Month-over-month and year-over-year fluctuation analysis.

Input DataFrames come from close_automation._load_single_period():
    [Account_Number, Account_Name, Balance]

Public API:
    analyze_flux(current_df, prior_month_df, prior_year_df=None, threshold_pct=5.0) -> dict
    export_flux_excel(result, output_path, ...) -> Path
"""

from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_NAVY      = "1F497D"
_BLUE_LT   = "EBF3FB"
_BLUE_MD   = "D9E1F2"
_GREEN     = "E2EFDA"
_GREEN_DK  = "375623"
_RED       = "FFC7CE"
_RED_DK    = "9C0006"
_YELLOW    = "FFEB9C"
_YELLOW_DK = "9C6500"
_GREY      = "F2F2F2"
_WHITE     = "FFFFFF"


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def analyze_flux(
    current_df: pd.DataFrame,
    prior_month_df: pd.DataFrame,
    prior_year_df: pd.DataFrame | None = None,
    threshold_pct: float = 5.0,
) -> dict:
    """Compute MoM (and optionally YoY) flux for each account.

    Returns dict with:
        accounts  -- full DataFrame (all accounts, all change columns)
        flagged   -- subset with any MoM or YoY flag
        summary   -- dict of high-level stats
        has_yoy   -- bool
    """
    cur = _normalize(current_df).rename(columns={"Balance": "Current"})
    pm  = _normalize(prior_month_df)[["Account_Number", "Balance"]].rename(
              columns={"Balance": "Prior_Month"})

    df = cur.merge(pm, on="Account_Number", how="outer")
    df["Current"]     = df["Current"].fillna(0)
    df["Prior_Month"] = df["Prior_Month"].fillna(0)
    if "Account_Name" not in df.columns:
        df["Account_Name"] = df["Account_Number"]
    df["Account_Name"] = df["Account_Name"].fillna(df["Account_Number"])

    df["MoM_Change"] = df["Current"] - df["Prior_Month"]
    df["MoM_Pct"]    = df.apply(lambda r: _safe_pct(r["MoM_Change"], r["Prior_Month"]), axis=1)
    df["MoM_Flag"]   = df["MoM_Pct"].abs() > threshold_pct
    df["MoM_Trend"]  = df["MoM_Change"].apply(lambda v: "up" if v > 0 else ("down" if v < 0 else "flat"))

    has_yoy = prior_year_df is not None
    if has_yoy:
        py = _normalize(prior_year_df)[["Account_Number", "Balance"]].rename(
                 columns={"Balance": "Prior_Year"})
        df = df.merge(py, on="Account_Number", how="left")
        df["Prior_Year"] = df["Prior_Year"].fillna(0)
        df["YoY_Change"] = df["Current"] - df["Prior_Year"]
        df["YoY_Pct"]    = df.apply(lambda r: _safe_pct(r["YoY_Change"], r["Prior_Year"]), axis=1)
        df["YoY_Flag"]   = df["YoY_Pct"].abs() > threshold_pct
        df["YoY_Trend"]  = df["YoY_Change"].apply(lambda v: "up" if v > 0 else ("down" if v < 0 else "flat"))

    df = df.sort_values("Account_Number").reset_index(drop=True)

    if has_yoy:
        flagged = df[df["MoM_Flag"] | df["YoY_Flag"]].copy()
    else:
        flagged = df[df["MoM_Flag"]].copy()

    mom_flagged = int(df["MoM_Flag"].sum())
    yoy_flagged = int(df["YoY_Flag"].sum()) if has_yoy else 0

    if not df.empty and df["MoM_Change"].abs().max() > 0:
        top_idx      = df["MoM_Change"].abs().idxmax()
        top_row      = df.loc[top_idx]
        top_mom_name = f"{top_row['Account_Number']} {top_row['Account_Name']}"
        top_mom_amt  = float(top_row["MoM_Change"])
    else:
        top_mom_name = None
        top_mom_amt  = 0.0

    return {
        "accounts": df,
        "flagged":  flagged,
        "has_yoy":  has_yoy,
        "summary": {
            "total_accounts":  len(df),
            "mom_flagged":     mom_flagged,
            "yoy_flagged":     yoy_flagged,
            "any_flagged":     len(flagged),
            "threshold_pct":   threshold_pct,
            "top_mom_account": top_mom_name,
            "top_mom_change":  top_mom_amt,
        },
    }


def export_flux_excel(
    result: dict,
    output_path,
    commentary=None,
    period_label: str = "",
    prior_month_label: str = "",
    prior_year_label: str = "",
    client_name: str = "",
) -> Path:
    """Write a multi-sheet flux analysis workbook.

    Sheets:
        Flux Analysis    -- full account list
        Material Fluxes  -- flagged accounts only
        AI Commentary    -- included when commentary is provided
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    df      = result["accounts"]
    flagged = result["flagged"]
    has_yoy = result["has_yoy"]

    wb = Workbook()

    ws = wb.active
    ws.title = "Flux Analysis"
    _write_flux_sheet(
        ws, df, has_yoy, client_name, period_label,
        prior_month_label, prior_year_label, result["summary"],
    )

    ws2 = wb.create_sheet("Material Fluxes")
    if flagged.empty:
        ws2.append(["No material fluxes detected above the threshold."])
        ws2["A1"].font = Font(name="Calibri", italic=True, color=_NAVY)
    else:
        _write_flux_sheet(
            ws2, flagged, has_yoy, client_name,
            f"Material Fluxes — {period_label}",
            prior_month_label, prior_year_label, result["summary"],
            flagged_only=True,
        )

    if commentary:
        ws3 = wb.create_sheet("AI Commentary")
        _write_commentary_sheet(ws3, commentary, has_yoy, client_name, period_label)

    wb.save(str(output_path))
    return output_path


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _normalize(df: pd.DataFrame) -> pd.DataFrame:
    col_map = {}
    for col in df.columns:
        low = col.lower().replace(" ", "_").replace("-", "_").replace(".", "_")
        if low in ("account_number", "acct_num", "account_no", "acct_no", "account"):
            col_map[col] = "Account_Number"
        elif low in ("account_name", "acct_name", "description", "name"):
            col_map[col] = "Account_Name"
        elif low in ("balance", "current_balance", "amount", "ending_balance",
                     "net_balance", "total"):
            col_map[col] = "Balance"
    return df.rename(columns=col_map).copy()


def _safe_pct(change: float, base: float) -> float:
    if base == 0:
        return 0.0 if change == 0 else (100.0 if change > 0 else -100.0)
    return round(change / abs(base) * 100, 2)


def _hdr_font(color=_WHITE):
    return Font(name="Calibri", bold=True, color=color, size=10)


def _cell_font(bold=False, color="000000"):
    return Font(name="Calibri", bold=bold, color=color, size=10)


def _fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)


def _border():
    thin = Side(style="thin", color="CCCCCC")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _write_flux_sheet(
    ws, df, has_yoy, client_name, current_label,
    prior_month_label, prior_year_label, summary, flagged_only=False,
):
    title = f"{client_name} — Flux Analysis" if client_name else "Flux Analysis"
    if current_label:
        title += f" — {current_label}"
    ws.append([title])
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=_NAVY)
    last_col = get_column_letter(13 if has_yoy else 8)
    ws.merge_cells(f"A1:{last_col}1")
    ws.row_dimensions[1].height = 22
    ws.append([])

    threshold = summary.get("threshold_pct", 5.0)
    ws.append([
        f"Variance threshold: {threshold:.1f}%",
        "",
        f"Total accounts: {summary.get('total_accounts', len(df))}",
        "",
        f"MoM flagged: {summary.get('mom_flagged', 0)}",
        "",
        f"YoY flagged: {summary.get('yoy_flagged', 0)}" if has_yoy else "",
    ])
    for cell in ws[3]:
        cell.font = Font(name="Calibri", size=9, color=_NAVY)
    ws.append([])

    if has_yoy:
        headers = [
            "Account #", "Account Name",
            f"Current ({current_label})" if current_label else "Current",
            f"Prior Month ({prior_month_label})" if prior_month_label else "Prior Month",
            "MoM Change", "MoM %", "Trend",
            f"Prior Year ({prior_year_label})" if prior_year_label else "Prior Year",
            "YoY Change", "YoY %", "Trend",
            "MoM", "YoY",
        ]
    else:
        headers = [
            "Account #", "Account Name",
            f"Current ({current_label})" if current_label else "Current",
            f"Prior Month ({prior_month_label})" if prior_month_label else "Prior Month",
            "MoM Change", "MoM %", "Trend", "Flag",
        ]

    ws.append(headers)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font      = _hdr_font()
        cell.fill      = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()
    ws.row_dimensions[hdr_row].height = 30

    for row_i, (_, row) in enumerate(df.iterrows()):
        mom_flag  = bool(row.get("MoM_Flag", False))
        mom_trend = "↑" if row["MoM_Trend"] == "up" else ("↓" if row["MoM_Trend"] == "down" else "—")

        if has_yoy:
            yoy_flag  = bool(row.get("YoY_Flag", False))
            yoy_trend = "↑" if row["YoY_Trend"] == "up" else ("↓" if row["YoY_Trend"] == "down" else "—")
            data = [
                str(row["Account_Number"]),
                str(row["Account_Name"]),
                round(float(row["Current"]),     2),
                round(float(row["Prior_Month"]), 2),
                round(float(row["MoM_Change"]),  2),
                round(float(row["MoM_Pct"]),     2),
                mom_trend,
                round(float(row.get("Prior_Year", 0)), 2),
                round(float(row.get("YoY_Change", 0)), 2),
                round(float(row.get("YoY_Pct",   0)), 2),
                yoy_trend,
                "●" if mom_flag else "",
                "●" if yoy_flag else "",
            ]
        else:
            yoy_flag = False
            data = [
                str(row["Account_Number"]),
                str(row["Account_Name"]),
                round(float(row["Current"]),     2),
                round(float(row["Prior_Month"]), 2),
                round(float(row["MoM_Change"]),  2),
                round(float(row["MoM_Pct"]),     2),
                mom_trend,
                "●" if mom_flag else "",
            ]

        ws.append(data)
        rn = ws.max_row

        base_fill = _fill(_GREY) if row_i % 2 == 1 else None
        if mom_flag:
            row_fill = _fill(_RED) if abs(float(row["MoM_Pct"])) > 15 else _fill(_YELLOW)
        elif has_yoy and yoy_flag:
            row_fill = _fill(_BLUE_LT)
        else:
            row_fill = base_fill

        for j, cell in enumerate(ws[rn]):
            cell.font      = _cell_font()
            cell.border    = _border()
            cell.alignment = Alignment(
                horizontal="left" if j == 1 else "center",
                vertical="center",
            )
            if row_fill:
                cell.fill = row_fill

        # Bold green/red on change and pct columns
        mom_chg_col = 5
        mom_pct_col = 6
        mom_color   = _RED_DK if float(row["MoM_Change"]) < 0 else _GREEN_DK
        for c in [mom_chg_col, mom_pct_col]:
            ws.cell(rn, c).font = Font(name="Calibri", size=10, bold=True, color=mom_color)
            if row_fill:
                ws.cell(rn, c).fill = row_fill

        if has_yoy:
            yoy_chg_col = 9
            yoy_pct_col = 10
            yoy_color   = _RED_DK if float(row.get("YoY_Change", 0)) < 0 else _GREEN_DK
            for c in [yoy_chg_col, yoy_pct_col]:
                ws.cell(rn, c).font = Font(name="Calibri", size=10, bold=True, color=yoy_color)
                if row_fill:
                    ws.cell(rn, c).fill = row_fill

        # Number formats
        for c in [3, 4, 5]:
            ws.cell(rn, c).number_format = "#,##0.00"
        ws.cell(rn, 6).number_format = '0.00"%"'
        if has_yoy:
            for c in [8, 9]:
                ws.cell(rn, c).number_format = "#,##0.00"
            ws.cell(rn, 10).number_format = '0.00"%"'

    # Column widths
    if has_yoy:
        widths = [12, 36, 16, 16, 14, 9, 6, 16, 14, 9, 6, 6, 6]
    else:
        widths = [12, 36, 16, 16, 14, 9, 6, 6]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = ws.cell(hdr_row + 1, 1)


def _write_commentary_sheet(ws, commentary, has_yoy, client_name, period_label):
    title = f"AI Flux Commentary — {period_label}" if period_label else "AI Flux Commentary"
    if client_name:
        title = f"{client_name} — {title}"
    ws.append([title])
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color=_NAVY)
    n_cols = 6 if has_yoy else 5
    ws.merge_cells(f"A1:{get_column_letter(n_cols)}1")
    ws.row_dimensions[1].height = 22
    ws.append([])

    if has_yoy:
        headers = ["Account #", "Account Name", "MoM %", "YoY %", "Analysis", "Recommended Action"]
    else:
        headers = ["Account #", "Account Name", "MoM %", "Analysis", "Recommended Action"]

    ws.append(headers)
    hdr_row = ws.max_row
    for cell in ws[hdr_row]:
        cell.font      = _hdr_font()
        cell.fill      = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border()
    ws.row_dimensions[hdr_row].height = 22

    for i, item in enumerate(commentary):
        mom_pct = item.get("mom_pct", "")
        if isinstance(mom_pct, float):
            mom_pct = round(mom_pct, 2)
        if has_yoy:
            yoy_pct = item.get("yoy_pct", "")
            if isinstance(yoy_pct, float):
                yoy_pct = round(yoy_pct, 2)
            row = [
                str(item.get("account_number", "")),
                str(item.get("account_name",   "")),
                mom_pct,
                yoy_pct,
                str(item.get("analysis", "")),
                str(item.get("action",   "")),
            ]
        else:
            row = [
                str(item.get("account_number", "")),
                str(item.get("account_name",   "")),
                mom_pct,
                str(item.get("analysis", "")),
                str(item.get("action",   "")),
            ]
        ws.append(row)
        rn = ws.max_row
        fill = _fill(_BLUE_LT) if i % 2 == 0 else _fill(_WHITE)
        n_center_cols = 4 if has_yoy else 3
        for j, cell in enumerate(ws[rn]):
            cell.font      = _cell_font()
            cell.fill      = fill
            cell.border    = _border()
            cell.alignment = Alignment(
                horizontal="center" if j < n_center_cols else "left",
                vertical="top",
                wrap_text=True,
            )
        ws.row_dimensions[rn].height = 52

    if has_yoy:
        widths = [12, 32, 9, 9, 55, 40]
    else:
        widths = [12, 32, 9, 55, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(hdr_row + 1, 1)
