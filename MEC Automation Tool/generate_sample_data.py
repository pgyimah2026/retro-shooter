"""generate_sample_data.py — Create realistic sample trial balance Excel files for testing.

Creates:
  data/ABC_Corp/trial_balance_2026-04.xlsx  (current month)
  data/ABC_Corp/trial_balance_2026-03.xlsx  (prior month)

Intentional variances (>5% threshold) in 6 accounts:
  1000 Cash                  +17.5%  (AR collections arrived)
  1100 Accounts Receivable   +15.0%  (delayed collections)
  4000 Product Sales Domestic +8.0%  (business growth)
  5000 Raw Materials Used    +12.0%  (tied to higher revenue)
  6300 Office Supplies      +500.0%  (one-time bulk order)
  6600 Marketing & Advertising -25.0% (campaign ended)
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

# ---------------------------------------------------------------------------
# Data — (account_number, account_name, apr_balance, mar_balance)
# All balances are absolute values (sign implied by account type)
# ---------------------------------------------------------------------------

ACCOUNTS = [
    # ── Assets ──────────────────────────────────────────────────────────────
    ("1000", "Cash and Cash Equivalents",           235_000,   200_000),   # +17.5% FLAGGED
    ("1010", "Petty Cash",                            2_500,     2_500),   #   0.0%
    ("1100", "Accounts Receivable",                 387_500,   337_000),   # +15.0% FLAGGED
    ("1110", "Allowance for Doubtful Accounts",      19_000,    18_500),   #  +2.7%
    ("1200", "Inventory — Raw Materials",            89_200,    86_400),   #  +3.2%
    ("1210", "Inventory — Work in Progress",         43_600,    42_800),   #  +1.9%
    ("1220", "Inventory — Finished Goods",           67_400,    65_700),   #  +2.6%
    ("1300", "Prepaid Expenses",                     18_500,    19_200),   #  -3.6%
    ("1400", "Other Current Assets",                  8_750,     8_900),   #  -1.7%
    ("1500", "Property, Plant & Equipment",       1_250_000, 1_250_000),   #   0.0%
    ("1510", "Accumulated Depreciation — PP&E",     387_500,   381_250),   #  +1.6%
    ("1600", "Intangible Assets",                    75_000,    75_000),   #   0.0%
    ("1610", "Accumulated Amortization",             22_500,    21_875),   #  +2.9%
    # ── Liabilities ─────────────────────────────────────────────────────────
    ("2000", "Accounts Payable",                    124_300,   121_400),   #  +2.4%
    ("2100", "Accrued Liabilities",                  38_500,    37_800),   #  +1.9%
    ("2110", "Accrued Payroll & Benefits",           67_800,    66_200),   #  +2.4%
    ("2200", "Current Portion of Long-Term Debt",    48_000,    48_000),   #   0.0%
    ("2300", "Deferred Revenue",                     31_200,    30_400),   #  +2.6%
    ("2400", "Other Current Liabilities",            12_400,    12_100),   #  +2.5%
    ("2500", "Long-Term Debt",                      425_000,   437_000),   #  -2.7%
    ("2600", "Deferred Tax Liability",               52_300,    51_200),   #  +2.1%
    # ── Equity ──────────────────────────────────────────────────────────────
    ("3000", "Common Stock",                        100_000,   100_000),   #   0.0%
    ("3100", "Additional Paid-in Capital",          350_000,   350_000),   #   0.0%
    ("3200", "Retained Earnings",                   687_500,   671_800),   #  +2.3%
    ("3300", "Treasury Stock",                       45_000,    45_000),   #   0.0%
    # ── Revenue ─────────────────────────────────────────────────────────────
    ("4000", "Product Sales — Domestic",            342_500,   317_100),   #  +8.0% FLAGGED
    ("4100", "Product Sales — International",       122_100,   119_200),   #  +2.4%
    ("4200", "Service Revenue",                      82_500,    80_900),   #  +2.0%
    ("4300", "Shipping & Handling Revenue",          13_450,    13_150),   #  +2.3%
    ("4900", "Sales Returns & Allowances",           18_500,    17_900),   #  +3.4%
    # ── Cost of Goods Sold ───────────────────────────────────────────────────
    ("5000", "Raw Materials Used",                   98_700,    88_100),   # +12.0% FLAGGED
    ("5100", "Direct Labor",                         62_000,    60_200),   #  +3.0%
    ("5200", "Manufacturing Overhead",               39_800,    38_600),   #  +3.1%
    ("5300", "Freight-In",                           11_700,    11_400),   #  +2.6%
    # ── Operating Expenses ───────────────────────────────────────────────────
    ("6000", "Salaries & Wages",                    122_500,   120_400),   #  +1.7%
    ("6010", "Payroll Taxes & Benefits",             24_500,    24_080),   #  +1.7%
    ("6100", "Rent Expense",                         28_500,    28_500),   #   0.0%
    ("6200", "Utilities",                             8_750,     8_900),   #  -1.7%
    ("6300", "Office Supplies & Equipment",          31_200,     5_200),   # +500.0% FLAGGED
    ("6400", "Insurance Expense",                    12_300,    12_300),   #   0.0%
    ("6500", "Depreciation & Amortization",           7_500,     7_500),   #   0.0%
    ("6600", "Marketing & Advertising",              18_750,    25_000),   # -25.0% FLAGGED
    ("6700", "Travel & Entertainment",                8_900,     8_700),   #  +2.3%
    ("6800", "Professional Fees",                    14_800,    14_500),   #  +2.1%
    ("6900", "Other Operating Expenses",              7_400,     7_200),   #  +2.8%
    # ── Other Income / Expense ───────────────────────────────────────────────
    ("7000", "Interest Income",                       3_200,     3_100),   #  +3.2%
    ("7200", "Other Income",                          4_500,     4_400),   #  +2.3%
    ("8000", "Interest Expense",                     18_500,    18_750),   #  -1.3%
    ("8100", "Loss on Sale of Assets",                    0,         0),   #   0.0%
    ("8200", "Income Tax Expense",                   27_500,    26_800),   #  +2.6%
]

# ---------------------------------------------------------------------------
# Styling constants
# ---------------------------------------------------------------------------

_NAVY  = "1F497D"
_LIGHT = "DCE6F1"
_WHITE = "FFFFFF"
_GRAY  = "F2F2F2"

_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_HEADER_FONT = Font(bold=True, color=_WHITE, name="Calibri", size=11)
_ALT_FILL    = PatternFill("solid", fgColor=_GRAY)
_TITLE_FONT  = Font(bold=True, name="Calibri", size=13)
_LABEL_FONT  = Font(name="Calibri", size=10)
_NUM_FMT     = '#,##0.00'

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(
    left=_THIN, right=_THIN, top=_THIN, bottom=_THIN,
)


def _write_tb(path: Path, period_label: str, balance_col_index: int) -> None:
    """Write one trial balance workbook.

    Row 1 = column headers (Account_Number, Account_Name, Balance) so the
    tool's _load_single_period() can read the file directly without skiprows.
    The sheet tab name carries the period label.

    Args:
        path:              Output file path.
        period_label:      e.g. "April 2026" — used as the sheet tab name.
        balance_col_index: 2 for April (index into each ACCOUNTS tuple),
                           3 for March.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = period_label

    # Header row — ROW 1 (no title row; tool reads row 1 as headers)
    headers = ["Account_Number", "Account_Name", "Balance"]
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.border    = _BORDER
        cell.alignment = Alignment(horizontal="center" if col != 2 else "left",
                                   vertical="center")
    ws.row_dimensions[1].height = 18

    # Data rows starting at row 2
    for i, row_data in enumerate(ACCOUNTS):
        row_num  = i + 2
        acct_num  = row_data[0]
        acct_name = row_data[1]
        balance   = row_data[balance_col_index]

        fill = _ALT_FILL if i % 2 == 1 else PatternFill(fill_type=None)

        c_num = ws.cell(row=row_num, column=1, value=acct_num)
        c_num.font          = Font(name="Calibri", size=10)
        c_num.fill          = fill
        c_num.border        = _BORDER
        c_num.number_format = "@"   # force text so account numbers stay as-is
        c_num.alignment     = Alignment(horizontal="center")

        c_name = ws.cell(row=row_num, column=2, value=acct_name)
        c_name.font      = Font(name="Calibri", size=10)
        c_name.fill      = fill
        c_name.border    = _BORDER
        c_name.alignment = Alignment(horizontal="left", indent=1)

        c_bal = ws.cell(row=row_num, column=3, value=balance)
        c_bal.font          = Font(name="Calibri", size=10)
        c_bal.fill          = fill
        c_bal.border        = _BORDER
        c_bal.number_format = _NUM_FMT
        c_bal.alignment     = Alignment(horizontal="right")

        ws.row_dimensions[row_num].height = 16

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18

    # Freeze header row
    ws.freeze_panes = "A2"

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    total_val = sum(r[balance_col_index] for r in ACCOUNTS)
    print(f"  Saved: {path}  ({len(ACCOUNTS)} accounts, total={total_val:,.2f})")


def main() -> None:
    base = Path(__file__).parent / "data" / "ABC_Corp"

    print("\nGenerating sample trial balance files…\n")

    _write_tb(base / "trial_balance_2026-04.xlsx", "April 2026",  balance_col_index=2)
    _write_tb(base / "trial_balance_2026-03.xlsx", "March 2026",  balance_col_index=3)

    print("\nDone.  Expected flagged variances (>5% threshold):")
    threshold = 5.0
    for acct_num, acct_name, apr, mar in ACCOUNTS:
        if mar == 0:
            continue
        pct = (apr - mar) / abs(mar) * 100
        if abs(pct) > threshold:
            direction = "[+]" if pct > 0 else "[-]"
            print(f"  {direction} {acct_num}  {acct_name:<40}  {pct:+.1f}%"
                  f"  (${mar:>10,.0f} -> ${apr:>10,.0f})")

    print()


if __name__ == "__main__":
    main()
