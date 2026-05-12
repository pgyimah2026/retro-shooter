"""generate_sample_flux_data.py -- Create prior-year trial balance for flux analysis testing.

Writes:
  data/ABC_Corp/trial_balance_2025-04.xlsx  -- April 2025 (prior year for YoY)

April 2025 tells a "pre-growth" story compared to April 2026:
  Revenue was ~12-15% lower (business was smaller)
  COGS was proportionally lower
  Marketing spend was ~30% higher (big campaign that year, since pulled back)
  Salaries were ~8% lower (fewer headcount)
  Cash was ~25% lower
  AR was ~20% lower
"""

from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# April 2025 balances -- same account structure as 2026-04 but scaled down / different mix
ACCOUNTS_2025_04 = [
    # ── Assets ────────────────────────────────────────────────────────────────
    ("1000", "Cash and Cash Equivalents",          178_000),   # -24.3% vs Apr 2026
    ("1010", "Petty Cash",                           2_500),   #  0%
    ("1100", "Accounts Receivable",                310_000),   # -20.0% vs Apr 2026
    ("1110", "Allowance for Doubtful Accounts",     16_200),   # -14.7%
    ("1200", "Inventory — Raw Materials",            82_500),   #  -7.5%
    ("1210", "Inventory — Work in Progress",         41_200),   #  -5.5%
    ("1220", "Inventory — Finished Goods",           63_800),   #  -5.3%
    ("1300", "Prepaid Expenses",                     17_800),   #  -3.8%
    ("1400", "Other Current Assets",                  8_400),   #  -4.0%
    ("1500", "Property, Plant & Equipment",       1_050_000),   # -16.0% (less capex)
    ("1510", "Accumulated Depreciation — PP&E",    312_500),   # -19.4%
    ("1600", "Intangible Assets",                    75_000),   #  0%
    ("1610", "Accumulated Amortization",             15_000),   # -33.3% (less amort)
    # ── Liabilities ───────────────────────────────────────────────────────────
    ("2000", "Accounts Payable",                   105_200),   # -15.4%
    ("2100", "Accrued Liabilities",                  34_500),   # -10.4%
    ("2110", "Accrued Payroll & Benefits",           61_000),   # -10.0%
    ("2200", "Current Portion of Long-Term Debt",    48_000),   #  0%
    ("2300", "Deferred Revenue",                     27_500),   # -11.9%
    ("2400", "Other Current Liabilities",            11_200),   #  -9.7%
    ("2500", "Long-Term Debt",                      490_000),   # +15.3% (more debt in 2025)
    ("2600", "Deferred Tax Liability",               46_800),   # -10.5%
    # ── Equity ────────────────────────────────────────────────────────────────
    ("3000", "Common Stock",                        100_000),   #  0%
    ("3100", "Additional Paid-in Capital",          350_000),   #  0%
    ("3200", "Retained Earnings",                   548_000),   # -20.2%
    ("3300", "Treasury Stock",                       45_000),   #  0%
    # ── Revenue ───────────────────────────────────────────────────────────────
    ("4000", "Product Sales — Domestic",            291_200),   # -15.0% vs Apr 2026
    ("4100", "Product Sales — International",       102_400),   # -16.1%
    ("4200", "Service Revenue",                      71_300),   # -13.6%
    ("4300", "Shipping & Handling Revenue",          11_800),   # -12.3%
    ("4900", "Sales Returns & Allowances",           16_200),   # -12.4%
    # ── Cost of Goods Sold ────────────────────────────────────────────────────
    ("5000", "Raw Materials Used",                   82_100),   # -16.8% vs Apr 2026
    ("5100", "Direct Labor",                         54_500),   # -12.1%
    ("5200", "Manufacturing Overhead",               35_200),   # -11.6%
    ("5300", "Freight-In",                           10_200),   # -12.8%
    # ── Operating Expenses ────────────────────────────────────────────────────
    ("6000", "Salaries & Wages",                    113_200),   #  -7.6% vs Apr 2026
    ("6010", "Payroll Taxes & Benefits",             22_600),   #  -7.8%
    ("6100", "Rent Expense",                         25_500),   # -10.5% (smaller space)
    ("6200", "Utilities",                             8_100),   #  -7.4%
    ("6300", "Office Supplies & Equipment",           4_800),   # -84.6% (vs Apr 2026 bulk order)
    ("6400", "Insurance Expense",                    11_200),   #  -8.9%
    ("6500", "Depreciation & Amortization",           6_250),   # -16.7%
    ("6600", "Marketing & Advertising",              32_500),   # +73.3% vs Apr 2026 -- big campaign
    ("6700", "Travel & Entertainment",                9_800),   # +10.1% vs Apr 2026
    ("6800", "Professional Fees",                    13_500),   #  -8.8%
    ("6900", "Other Operating Expenses",              6_800),   #  -8.1%
    # ── Other Income / Expense ────────────────────────────────────────────────
    ("7000", "Interest Income",                       2_400),   # -25.0% vs Apr 2026
    ("7200", "Other Income",                          3_900),   # -13.3%
    ("8000", "Interest Expense",                     21_500),   # +16.2% vs Apr 2026 (higher debt)
    ("8100", "Loss on Sale of Assets",                    0),   #  0%
    ("8200", "Income Tax Expense",                   22_100),   # -19.6%
]

_NAVY  = "1F497D"
_GRAY  = "F2F2F2"
_WHITE = "FFFFFF"

_HEADER_FILL = PatternFill("solid", fgColor=_NAVY)
_HEADER_FONT = Font(bold=True, color=_WHITE, name="Calibri", size=11)
_ALT_FILL    = PatternFill("solid", fgColor=_GRAY)
_BORDER      = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)


def main():
    out_dir = Path(__file__).parent / "data" / "ABC_Corp"
    out_dir.mkdir(parents=True, exist_ok=True)

    path   = out_dir / "trial_balance_2025-04.xlsx"
    label  = "April 2025"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = label

    headers = ["Account_Number", "Account_Name", "Balance"]
    for col, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=hdr)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.border    = _BORDER
        cell.alignment = Alignment(horizontal="center" if col != 2 else "left",
                                   vertical="center")
    ws.row_dimensions[1].height = 18

    for i, (acct_num, acct_name, balance) in enumerate(ACCOUNTS_2025_04):
        rn   = i + 2
        fill = _ALT_FILL if i % 2 == 1 else PatternFill(fill_type=None)

        c_num = ws.cell(row=rn, column=1, value=acct_num)
        c_num.font          = Font(name="Calibri", size=10)
        c_num.fill          = fill
        c_num.border        = _BORDER
        c_num.number_format = "@"
        c_num.alignment     = Alignment(horizontal="center")

        c_name = ws.cell(row=rn, column=2, value=acct_name)
        c_name.font      = Font(name="Calibri", size=10)
        c_name.fill      = fill
        c_name.border    = _BORDER
        c_name.alignment = Alignment(horizontal="left", indent=1)

        c_bal = ws.cell(row=rn, column=3, value=balance)
        c_bal.font          = Font(name="Calibri", size=10)
        c_bal.fill          = fill
        c_bal.border        = _BORDER
        c_bal.number_format = "#,##0.00"
        c_bal.alignment     = Alignment(horizontal="right")

        ws.row_dimensions[rn].height = 16

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 42
    ws.column_dimensions["C"].width = 18
    ws.freeze_panes = "A2"

    wb.save(str(path))
    print(f"Saved: {path}  ({len(ACCOUNTS_2025_04)} accounts)")
    print()
    print("YoY flux highlights (April 2026 vs April 2025):")
    from generate_sample_data import ACCOUNTS as ACCOUNTS_2026
    acct_map = {a[0]: a for a in ACCOUNTS_2026}
    prev_map = {a[0]: a[2] for a in ACCOUNTS_2025_04}
    for acct_num, acct_name, bal_2026, _ in ACCOUNTS_2026:
        bal_2025 = prev_map.get(acct_num, 0)
        if bal_2025 == 0:
            continue
        pct = (bal_2026 - bal_2025) / abs(bal_2025) * 100
        if abs(pct) > 10:
            direction = "[+]" if pct > 0 else "[-]"
            print(f"  {direction} {acct_num}  {acct_name:<40}  {pct:+.1f}%"
                  f"  (${bal_2025:>10,.0f} -> ${bal_2026:>10,.0f})")


if __name__ == "__main__":
    main()
