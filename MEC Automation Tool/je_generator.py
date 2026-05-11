"""je_generator.py — Create an Excel journal entry template from flagged variances.

Public API:
    from je_generator import create_je_template

    path = create_je_template(
        flagged=flagged_df,      # from variance_calculator (Flagged==True rows)
        config=client_config,    # from load_config
        period_label="April 2026",
        output_dir="output",
        commentary=[...],        # optional, from commentary_generator
        full_report=df,          # optional, enriches account dropdown labels
    )
    print(path)

CLI (standalone sanity check):
    python je_generator.py --period "April 2026" --output output
"""

import argparse
import calendar
import logging
import re
from datetime import date
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Layout constants
# ---------------------------------------------------------------------------

# Row indices (1-based)
_ROW_TITLE  = 1   # merged title band
_ROW_HEADER = 2   # column headers
_ROW_FIRST  = 3   # first data row
_MAX_ROWS   = 500 # validation / CF applied over this many data rows
_ROW_LAST   = _ROW_FIRST + _MAX_ROWS - 1   # = 502

# Column indices and metadata: (label, width in chars)
_JE_COLS = [
    ("JE #",          8),   # A
    ("Date",         12),   # B
    ("Account #",    12),   # C  ← dropdown + numeric validation
    ("Account Name", 24),   # D
    ("Description",  38),   # E
    ("Debit ($)",    14),   # F  ← numeric validation
    ("Credit ($)",   14),   # G  ← numeric validation
    ("Memo / Notes", 32),   # H
]
_N_COLS = len(_JE_COLS)

# ---------------------------------------------------------------------------
# Style constants
# ---------------------------------------------------------------------------

_FONT_NAME = "Calibri"

_FILL_HDR     = PatternFill("solid", fgColor="1F497D")   # navy header
_FILL_ODD_JE  = PatternFill("solid", fgColor="FFFFFF")   # white row
_FILL_EVEN_JE = PatternFill("solid", fgColor="F2F2F2")   # light grey row
_FILL_OOB     = PatternFill("solid", fgColor="FFC7CE")   # red — out of balance
_FILL_KPI_LBL = PatternFill("solid", fgColor="D9E1F2")   # summary label cells
_FILL_KPI_VAL = PatternFill("solid", fgColor="EBF3FB")   # summary value cells
_FILL_BAL     = PatternFill("solid", fgColor="E2EFDA")   # balanced JE (green)
_FILL_UNBP    = PatternFill("solid", fgColor="FFC7CE")   # unbalanced JE (red)

_FONT_HDR     = Font(name=_FONT_NAME, bold=True, color="FFFFFF", size=10)
_FONT_BODY    = Font(name=_FONT_NAME, size=10)
_FONT_BOLD    = Font(name=_FONT_NAME, bold=True, size=10)
_FONT_OOB     = Font(name=_FONT_NAME, bold=True, color="9C0006", size=10)
_FONT_TITLE   = Font(name=_FONT_NAME, bold=True, color="1F497D", size=13)
_FONT_GOOD    = Font(name=_FONT_NAME, bold=True, color="375623", size=11)
_FONT_BAD     = Font(name=_FONT_NAME, bold=True, color="9C0006", size=11)
_FONT_NOTE    = Font(name=_FONT_NAME, italic=True, color="666666", size=8)

_THIN_SIDE    = Side(style="thin", color="BFBFBF")
_BORDER_THIN  = Border(
    left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE
)

_ALIGN_C  = Alignment(horizontal="center", vertical="center")
_ALIGN_R  = Alignment(horizontal="right",  vertical="center")
_ALIGN_L  = Alignment(horizontal="left",   vertical="center")
_ALIGN_W  = Alignment(horizontal="left",   vertical="center", wrap_text=True)

_FMT_ACCT = '#,##0.00_);[Red](#,##0.00)'
_FMT_DATE = 'YYYY-MM-DD'


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def create_je_template(
    flagged: pd.DataFrame,
    config: dict,
    period_label: str,
    output_dir: str = "output",
    commentary: Optional[list] = None,
    full_report: Optional[pd.DataFrame] = None,
) -> Path:
    """Create an Excel journal entry template for each flagged variance.

    Workbook structure:
        Journal Entries  — main template with pre-populated JE pairs, data
                           validation (account dropdown, numeric amounts), and
                           conditional formatting for out-of-balance detection.
        Summary          — total debits, credits, out-of-balance amount, and a
                           per-JE balance status table.
        Account_List     — hidden sheet that feeds the account dropdown.

    Each flagged variance produces two pre-populated rows: a debit line for the
    flagged account and a blank credit line (offset account TBD by the preparer).
    The preparer may add additional rows and the balance check updates live.

    Args:
        flagged:      DataFrame of flagged variance rows. Expected columns:
                      Account_Number, Account_Name, Variance_Amount,
                      Variance_Pct. Rows where Flagged==False are ignored
                      if the full report is passed.
        config:       Client configuration dict. Keys used: client_name,
                      account_groups (dict mapping group label → [acct nums]).
        period_label: Human-readable period, e.g. "April 2026". Determines
                      the pre-filled entry date (last day of the month).
        output_dir:   Directory where the .xlsx file is written.
        commentary:   Optional list of dicts from commentary_generator, each
                      with account_number and commentary. Used to pre-fill the
                      Description field with the first sentence.
        full_report:  Optional full DataFrame (all accounts). When provided,
                      account names are sourced from it to label the dropdown.

    Returns:
        Path to the saved .xlsx file.

    Raises:
        OSError: If the output file cannot be written.
    """
    commentary = commentary or []
    commentary_map = {c["account_number"]: c["commentary"] for c in commentary}

    # Accept either pre-filtered flagged rows or full report (filter internally)
    if "Flagged" in flagged.columns:
        data = flagged[flagged["Flagged"]].copy().reset_index(drop=True)
    else:
        data = flagged.copy().reset_index(drop=True)

    period_date  = _parse_period(period_label)
    account_list = _build_account_list(config, full_report)
    je_rows      = _build_je_rows(data, period_date, commentary_map)

    wb             = Workbook()
    ws_je          = wb.active
    ws_je.title    = "Journal Entries"
    ws_sum         = wb.create_sheet("Summary")
    ws_acct        = wb.create_sheet("Account_List")
    ws_acct.sheet_state = "hidden"

    _write_account_list_sheet(ws_acct, account_list)
    _write_je_sheet(ws_je, je_rows, account_list, period_date, config, period_label)
    _write_summary_sheet(ws_sum, je_rows, config, period_label)

    client_slug = re.sub(r"[^A-Za-z0-9]+", "_", config.get("client_name", "client")).strip("_")
    period_slug = re.sub(r"[^A-Za-z0-9]+", "_", period_label).strip("_")
    filename    = f"journal_entries_template_{client_slug}_{period_slug}.xlsx"

    out_path = Path(output_dir) / filename
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out_path))

    n_jes = len(data)
    log.info(
        "JE template saved: '%s' (%d JE(s), %d rows, period date %s).",
        out_path, n_jes, len(je_rows), period_date,
    )
    return out_path


# ---------------------------------------------------------------------------
# Data builders
# ---------------------------------------------------------------------------

def _parse_period(period_label: str) -> date:
    """Parse 'Month YYYY' into the last calendar day of that month.

    Falls back to today's date when the label cannot be parsed.

    Args:
        period_label: String like "April 2026".

    Returns:
        date object representing the last day of the named month.
    """
    match = re.search(r"(\w+)\s+(\d{4})", period_label)
    if match:
        try:
            month_num = list(calendar.month_name).index(match.group(1).capitalize())
            year      = int(match.group(2))
            last_day  = calendar.monthrange(year, month_num)[1]
            return date(year, month_num, last_day)
        except (ValueError, IndexError):
            pass
    log.warning("Could not parse period '%s'; using today's date.", period_label)
    return date.today()


def _build_account_list(
    config: dict,
    full_report: Optional[pd.DataFrame],
) -> list:
    """Build an ordered list of (account_number, display_label) tuples.

    Account numbers are drawn from config['account_groups']. If full_report
    is provided, each number is matched to its Account_Name for a richer
    dropdown label. Duplicates across groups are deduplicated in insertion order.

    Args:
        config:      Client config dict.
        full_report: Optional full trial balance DataFrame.

    Returns:
        List of (str, str) tuples: (account_number, display_label).
    """
    name_map: dict = {}
    if full_report is not None and "Account_Number" in full_report.columns:
        name_map = dict(zip(
            full_report["Account_Number"].astype(str),
            full_report["Account_Name"].astype(str),
        ))

    accounts: list = []
    seen: set      = set()
    for group_accounts in config.get("account_groups", {}).values():
        for raw in group_accounts:
            num = str(raw).strip()
            if num in seen:
                continue
            seen.add(num)
            name  = name_map.get(num, "")
            label = f"{num} – {name}" if name else num
            accounts.append((num, label))

    return accounts


def _build_je_rows(
    data: pd.DataFrame,
    period_date: date,
    commentary_map: dict,
) -> list:
    """Build the list of JE row dicts (two rows per flagged variance).

    The first row (debit line) is pre-populated with the flagged account and
    the full variance amount in the Debit column. The second row (credit line)
    has the same amount in the Credit column and a blank Account # for the
    preparer to supply the offsetting account.

    Args:
        data:            Filtered DataFrame of flagged variances.
        period_date:     Pre-filled entry date (last day of the period).
        commentary_map:  Dict of {account_number: commentary_text}.

    Returns:
        List of dicts, each representing one spreadsheet row.
    """
    rows: list = []
    for je_num, (_, var) in enumerate(data.iterrows(), start=1):
        acct_num  = str(var["Account_Number"])
        acct_name = str(var["Account_Name"])
        var_amt   = float(var["Variance_Amount"])
        amount    = abs(var_amt)
        var_pct   = float(var.get("Variance_Pct", 0.0))

        # First sentence of AI commentary → description; fall back to generic
        raw_commentary = commentary_map.get(acct_num, "")
        if raw_commentary:
            first = re.split(r"(?<=[.!?])\s", raw_commentary.strip())[0]
            description = first[:120]
        else:
            direction   = "increase" if var_amt > 0 else "decrease"
            description = f"Adj – {acct_name} {direction} ({var_pct:+.1f}%)"

        memo = f"Variance: {var_pct:+.1f}%  |  Ref acct: {acct_num}"

        # Debit line (flagged account, preparer verifies DR vs CR direction)
        rows.append({
            "je_num":      je_num,
            "date":        period_date,
            "account_num": acct_num,
            "account_name":acct_name,
            "description": description,
            "debit":       amount,
            "credit":      None,
            "memo":        memo,
            "line_type":   "DR",
        })

        # Credit line (offset account — left blank for preparer)
        rows.append({
            "je_num":      je_num,
            "date":        period_date,
            "account_num": "",
            "account_name":"",
            "description": description,
            "debit":       None,
            "credit":      amount,
            "memo":        "",
            "line_type":   "CR",
        })

    return rows


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_account_list_sheet(ws, account_list: list) -> None:
    """Populate the hidden Account_List sheet used by the account dropdown.

    Args:
        ws:           openpyxl Worksheet object (hidden).
        account_list: List of (account_number, display_label) tuples.
    """
    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 35
    for i, (num, label) in enumerate(account_list, start=1):
        ws.cell(row=i, column=1, value=num)
        ws.cell(row=i, column=2, value=label)


def _write_je_sheet(
    ws,
    je_rows: list,
    account_list: list,
    period_date: date,
    config: dict,
    period_label: str,
) -> None:
    """Build the main Journal Entries worksheet.

    Args:
        ws:           openpyxl Worksheet object.
        je_rows:      List of row dicts from _build_je_rows.
        account_list: Account list for dropdown validation.
        period_date:  Pre-filled entry date.
        config:       Client config dict.
        period_label: Human-readable period string.
    """
    client = config.get("client_name", "")

    # ── Title row ────────────────────────────────────────────────────────────
    last_col_letter = get_column_letter(_N_COLS)
    ws.merge_cells(f"A{_ROW_TITLE}:{last_col_letter}{_ROW_TITLE}")
    tc = ws.cell(row=_ROW_TITLE, column=1)
    tc.value = (
        f"Journal Entry Template  ·  {client}  ·  Period: {period_label}  "
        f"·  Entry Date: {period_date.strftime('%B %d, %Y')}"
    )
    tc.font      = _FONT_TITLE
    tc.alignment = _ALIGN_L
    ws.row_dimensions[_ROW_TITLE].height = 22

    # ── Column headers ────────────────────────────────────────────────────────
    for ci, (label, width) in enumerate(_JE_COLS, start=1):
        cell = ws.cell(row=_ROW_HEADER, column=ci, value=label)
        cell.fill      = _FILL_HDR
        cell.font      = _FONT_HDR
        cell.alignment = _ALIGN_C
        cell.border    = _BORDER_THIN
        ws.column_dimensions[get_column_letter(ci)].width = width
    ws.row_dimensions[_ROW_HEADER].height = 18

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, row_data in enumerate(je_rows, start=_ROW_FIRST):
        _write_data_row(ws, ri, row_data)

    # ── Legend (below pre-populated data, above empty template space) ─────────
    legend_row = _ROW_FIRST + len(je_rows) + 1
    _write_legend(ws, legend_row)

    # ── Validation ────────────────────────────────────────────────────────────
    _apply_account_dropdown(ws, account_list)
    _apply_amount_validation(ws)

    # ── Conditional formatting — unbalanced JE highlight ──────────────────────
    _apply_balance_cf(ws)

    # ── Freeze and display settings ───────────────────────────────────────────
    ws.freeze_panes    = f"A{_ROW_FIRST}"
    ws.sheet_view.showGridLines = True


def _write_data_row(ws, row_idx: int, row_data: dict) -> None:
    """Write one JE row (debit or credit line) with fill and number formats.

    Args:
        ws:       openpyxl Worksheet object.
        row_idx:  1-based row index.
        row_data: Dict produced by _build_je_rows.
    """
    je_num = row_data["je_num"]
    fill   = _FILL_ODD_JE if je_num % 2 == 1 else _FILL_EVEN_JE

    col_values = [
        (je_num,                   _ALIGN_C,  _FMT_DATE,  False),
        (row_data["date"],         _ALIGN_C,  _FMT_DATE,  False),
        (row_data["account_num"],  _ALIGN_C,  None,       False),
        (row_data["account_name"], _ALIGN_L,  None,       False),
        (row_data["description"],  _ALIGN_W,  None,       False),
        (row_data["debit"],        _ALIGN_R,  _FMT_ACCT,  False),
        (row_data["credit"],       _ALIGN_R,  _FMT_ACCT,  False),
        (row_data["memo"],         _ALIGN_L,  None,       False),
    ]
    # JE # cell: plain integer, no date format
    col_values[0] = (je_num, _ALIGN_C, None, True)

    for ci, (value, alignment, num_fmt, bold) in enumerate(col_values, start=1):
        cell            = ws.cell(row=row_idx, column=ci, value=value)
        cell.fill       = fill
        cell.font       = Font(name=_FONT_NAME, size=10, bold=bold)
        cell.alignment  = alignment
        cell.border     = _BORDER_THIN
        if num_fmt:
            cell.number_format = num_fmt

    ws.row_dimensions[row_idx].height = 16


def _write_legend(ws, row_idx: int) -> None:
    """Write a compact colour legend and usage note below the data.

    Args:
        ws:      openpyxl Worksheet object.
        row_idx: Row to start the legend on.
    """
    ws.merge_cells(f"A{row_idx}:{get_column_letter(_N_COLS)}{row_idx}")
    cell = ws.cell(row=row_idx, column=1)
    cell.value = (
        "Legend:  Red row = JE out of balance (Debits ≠ Credits).  "
        "Alternating white/grey pairs = one JE.  "
        "Add new JEs below — validation and balance-check apply to all rows."
    )
    cell.font      = _FONT_NOTE
    cell.alignment = _ALIGN_L
    ws.row_dimensions[row_idx].height = 14


def _write_summary_sheet(
    ws,
    je_rows: list,
    config: dict,
    period_label: str,
) -> None:
    """Build the Summary worksheet with aggregate KPIs and a per-JE table.

    Args:
        ws:           openpyxl Worksheet object.
        je_rows:      Full JE row list from _build_je_rows.
        config:       Client config dict.
        period_label: Human-readable period string.
    """
    # Column widths
    for col, width in zip("ABCDEF", [26, 16, 26, 14, 14, 14]):
        ws.column_dimensions[col].width = width

    # ── Title ─────────────────────────────────────────────────────────────────
    ws.merge_cells("A1:F1")
    tc = ws.cell(row=1, column=1)
    tc.value     = f"Journal Entry Summary  ·  {config.get('client_name','')}  ·  {period_label}"
    tc.font      = _FONT_TITLE
    tc.alignment = _ALIGN_L
    ws.row_dimensions[1].height = 22

    # ── Aggregate KPIs (rows 3-8) ─────────────────────────────────────────────
    total_dr  = sum(r["debit"]  or 0 for r in je_rows)
    total_cr  = sum(r["credit"] or 0 for r in je_rows)
    oob       = round(total_dr - total_cr, 2)
    balanced  = oob == 0
    n_jes     = max((r["je_num"] for r in je_rows), default=0)

    kpis = [
        ("Total JEs Prepared",   str(n_jes),             None),
        ("Total Debit Lines",    str(sum(1 for r in je_rows if r["line_type"] == "DR")), None),
        ("Total Debits ($)",     total_dr,                _FMT_ACCT),
        ("Total Credits ($)",    total_cr,                _FMT_ACCT),
        ("Out-of-Balance ($)",   abs(oob),                _FMT_ACCT),
        ("Balance Status",       "BALANCED ✓" if balanced else "OUT OF BALANCE ✗", None),
    ]

    for ri, (label, value, num_fmt) in enumerate(kpis, start=3):
        lc = ws.cell(row=ri, column=1, value=label)
        lc.font      = _FONT_BOLD
        lc.fill      = _FILL_KPI_LBL
        lc.alignment = _ALIGN_L
        lc.border    = _BORDER_THIN

        vc = ws.cell(row=ri, column=2, value=value)
        vc.fill      = _FILL_KPI_VAL
        vc.alignment = _ALIGN_R
        vc.border    = _BORDER_THIN
        if num_fmt:
            vc.number_format = num_fmt
        if label == "Balance Status":
            vc.font = _FONT_GOOD if balanced else _FONT_BAD
        else:
            vc.font = _FONT_BODY

        ws.row_dimensions[ri].height = 18

    # ── Per-JE breakdown table ────────────────────────────────────────────────
    _write_je_breakdown(ws, je_rows, start_row=11)

    ws.freeze_panes = "A2"


def _write_je_breakdown(ws, je_rows: list, start_row: int) -> None:
    """Write the per-JE balance status table in the Summary sheet.

    Args:
        ws:        openpyxl Worksheet object.
        je_rows:   Full JE row list.
        start_row: Row index to start the table.
    """
    # Aggregate by JE number
    je_totals: dict = {}
    for r in je_rows:
        n = r["je_num"]
        if n not in je_totals:
            je_totals[n] = {
                "debit": 0.0, "credit": 0.0,
                "ref_acct": "", "description": "",
            }
        je_totals[n]["debit"]  += r["debit"]  or 0
        je_totals[n]["credit"] += r["credit"] or 0
        if r["line_type"] == "DR":                 # grab metadata from debit line
            je_totals[n]["ref_acct"]    = r["account_num"]
            je_totals[n]["description"] = r["description"]

    # Section heading
    ws.merge_cells(f"A{start_row}:F{start_row}")
    sh = ws.cell(row=start_row, column=1, value="Per-JE Balance Check")
    sh.font      = Font(name=_FONT_NAME, bold=True, color="1F497D", size=11)
    sh.alignment = _ALIGN_L
    ws.row_dimensions[start_row].height = 18

    # Table header
    hdr_row = start_row + 1
    headers = ["JE #", "Ref Account #", "Description", "Debit ($)", "Credit ($)", "Difference ($)"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=hdr_row, column=ci, value=h)
        c.fill      = _FILL_HDR
        c.font      = _FONT_HDR
        c.alignment = _ALIGN_C
        c.border    = _BORDER_THIN
    ws.row_dimensions[hdr_row].height = 16

    # Data rows
    for i, (je_num, totals) in enumerate(sorted(je_totals.items()), start=1):
        ri   = hdr_row + i
        diff = round(totals["debit"] - totals["credit"], 2)
        ok   = diff == 0
        fill = _FILL_BAL if ok else _FILL_UNBP
        font = _FONT_BODY if ok else Font(name=_FONT_NAME, bold=True, color="9C0006", size=10)

        row_vals = [
            (je_num,                    _ALIGN_C,  None),
            (totals["ref_acct"],        _ALIGN_C,  None),
            (totals["description"],     _ALIGN_W,  None),
            (totals["debit"],           _ALIGN_R,  _FMT_ACCT),
            (totals["credit"],          _ALIGN_R,  _FMT_ACCT),
            (diff,                      _ALIGN_R,  _FMT_ACCT),
        ]
        for ci, (val, align, fmt) in enumerate(row_vals, start=1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill      = fill
            c.font      = font
            c.alignment = align
            c.border    = _BORDER_THIN
            if fmt:
                c.number_format = fmt
        ws.row_dimensions[ri].height = 16

    # Totals footer
    footer_row = hdr_row + len(je_totals) + 1
    for ci, (label, value, fmt) in enumerate([
        ("TOTAL", "",
         None),
        ("", "", None),
        ("", "", None),
        ("", sum(t["debit"]  for t in je_totals.values()), _FMT_ACCT),
        ("", sum(t["credit"] for t in je_totals.values()), _FMT_ACCT),
        ("", round(sum(t["debit"] - t["credit"] for t in je_totals.values()), 2), _FMT_ACCT),
    ], start=1):
        # Use label for col 1, value for the rest
        content = label if ci == 1 else value
        c = ws.cell(row=footer_row, column=ci, value=content)
        c.fill      = _FILL_KPI_LBL
        c.font      = _FONT_BOLD
        c.alignment = _ALIGN_R if ci >= 4 else _ALIGN_L
        c.border    = _BORDER_THIN
        if fmt and value != "":
            c.number_format = fmt
    ws.row_dimensions[footer_row].height = 16


# ---------------------------------------------------------------------------
# Validation and conditional formatting helpers
# ---------------------------------------------------------------------------

def _apply_account_dropdown(ws, account_list: list) -> None:
    """Add an account number dropdown to the Account # column (C).

    Uses the hidden Account_List sheet for account lists with more than 20
    entries (avoids the 255-character inline formula limit). Shorter lists are
    inlined for portability.

    Args:
        ws:           Journal Entries worksheet.
        account_list: List of (account_number, label) tuples.
    """
    n   = len(account_list)
    col = "C"
    rng = f"{col}{_ROW_FIRST}:{col}{_ROW_LAST}"

    if n == 0:
        return

    if n <= 20:
        nums    = ",".join(a[0] for a in account_list)
        formula = f'"{nums}"'
    else:
        formula = f"'Account_List'!$A$1:$A${n}"

    dv = DataValidation(
        type="list",
        formula1=formula,
        allow_blank=True,
        showDropDown=False,        # False = SHOW the dropdown arrow (Excel XML quirk)
        showErrorMessage=True,
        errorStyle="warning",      # warning allows manual entries outside the list
        errorTitle="Account not in list",
        error="This account is not in the pre-defined list. Verify before saving.",
    )
    dv.sqref = rng
    ws.add_data_validation(dv)


def _apply_amount_validation(ws) -> None:
    """Restrict Debit (F) and Credit (G) columns to non-negative decimals.

    Args:
        ws: Journal Entries worksheet.
    """
    for col in ("F", "G"):
        rng = f"{col}{_ROW_FIRST}:{col}{_ROW_LAST}"
        dv  = DataValidation(
            type="decimal",
            operator="greaterThanOrEqual",
            formula1="0",
            allow_blank=True,
            showErrorMessage=True,
            errorStyle="warning",
            errorTitle="Invalid Amount",
            error="Debit/Credit amounts must be zero or positive. Use separate rows for adjustments.",
        )
        dv.sqref = rng
        ws.add_data_validation(dv)


def _apply_balance_cf(ws) -> None:
    """Highlight all rows of an out-of-balance JE in red.

    Uses a SUMIF formula to aggregate debits and credits by JE number. Any row
    whose JE number's debits ≠ credits is coloured with the OOB fill. The rule
    applies only to rows that have a JE number (skips blank rows).

    SUMIF logic:
        total_DR = SUMIF(JE_col, this_row_JE, Debit_col)
        total_CR = SUMIF(JE_col, this_row_JE, Credit_col)
        flag     = ROUND(total_DR - total_CR, 2) <> 0

    Args:
        ws: Journal Entries worksheet.
    """
    data_range = f"A{_ROW_FIRST}:{get_column_letter(_N_COLS)}{_ROW_LAST}"

    a_rng = f"$A${_ROW_FIRST}:$A${_ROW_LAST}"
    f_rng = f"$F${_ROW_FIRST}:$F${_ROW_LAST}"
    g_rng = f"$G${_ROW_FIRST}:$G${_ROW_LAST}"

    formula = (
        f'AND($A{_ROW_FIRST}<>"",'
        f'ROUND('
        f'SUMIF({a_rng},$A{_ROW_FIRST},{f_rng})'
        f'-SUMIF({a_rng},$A{_ROW_FIRST},{g_rng})'
        f',2)<>0)'
    )

    rule = FormulaRule(formula=[formula], fill=_FILL_OOB, font=_FONT_OOB)
    ws.conditional_formatting.add(data_range, rule)


# ---------------------------------------------------------------------------
# Standalone execution
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    logging.basicConfig(level=logging.DEBUG, format="%(levelname)-8s %(name)s: %(message)s")

    parser = argparse.ArgumentParser(description="Generate a JE Excel template.")
    parser.add_argument("--period", default="April 2026", help="Period label, e.g. 'April 2026'")
    parser.add_argument("--output", default="output",    help="Output directory")
    args = parser.parse_args()

    sample_flagged = pd.DataFrame({
        "Account_Number":  ["4000",         "5000",      "6000",      "7000"],
        "Account_Name":    ["Product Sales", "Materials", "Salaries",  "Interest Income"],
        "Current_Balance": [178_000,         72_000,      42_000,      1_200],
        "Prior_Balance":   [150_000,         60_000,      35_000,      1_000],
        "Variance_Amount": [ 28_000,         12_000,       7_000,        200],
        "Variance_Pct":    [  18.67,          20.00,       20.00,       20.00],
        "Flagged":         [   True,           True,        True,        True],
    })

    sample_config = {
        "client_name": "Acme Corporation",
        "variance_threshold_pct": 10,
        "variance_threshold_abs": 5000,
        "account_groups": {
            "Revenue":             ["4000", "4100", "4200", "4300"],
            "Cost of Goods Sold":  ["5000", "5100", "5200"],
            "Operating Expenses":  ["6000", "6100", "6200", "6300", "6400"],
            "Other Income/Expense":["7000", "7100", "8000"],
        },
    }

    sample_commentary = [
        {
            "account_number": "4000",
            "commentary": (
                "Product Sales increased $28,000 (18.7%) driven by strong Q2 seasonal demand "
                "and completion of a deferred Q1 contract. Review with revenue recognition policy."
            ),
        },
        {
            "account_number": "5000",
            "commentary": (
                "Materials expense rose $12,000 (20.0%) in line with higher production volume "
                "and an 8% supplier price increase effective April 1."
            ),
        },
        {
            "account_number": "6000",
            "commentary": (
                "Salaries increased $7,000 (20.0%) due to annual merit adjustments and a "
                "new part-time contractor added in April."
            ),
        },
        {
            "account_number": "7000",
            "commentary": (
                "Interest Income rose $200 (20.0%) from higher money-market yields. "
                "Expected to continue through Q3."
            ),
        },
    ]

    out = create_je_template(
        flagged=sample_flagged,
        config=sample_config,
        period_label=args.period,
        output_dir=args.output,
        commentary=sample_commentary,
    )
    print(f"Template saved: {out}")
