"""variance_calculator.py — Compute account-level variances and export to Excel.

Two entry points depending on your data shape:

    # Two separate DataFrames, each with a balance column:
    from variance_calculator import calculate_variances, export_to_excel

    report = calculate_variances(current_df, prior_df)
    export_to_excel(report, "output/variance_report.xlsx")

    # One DataFrame that already has Current_Balance and Prior_Balance columns
    # (e.g. output from data_parser.load_trial_balance):
    from variance_calculator import calculate_variances_single, export_to_excel

    report = calculate_variances_single(df)
    export_to_excel(report, "output/variance_report.xlsx")
"""

import logging
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

# Column names used throughout this module
_OUT_COLS = [
    "Account_Number",
    "Account_Name",
    "Current_Balance",
    "Prior_Balance",
    "Variance_Amount",
    "Variance_Pct",
    "Flagged",
]

# Excel styling constants
_FILL_HEADER = PatternFill("solid", fgColor="1F497D")
_FILL_FLAGGED = PatternFill("solid", fgColor="FFCCCC")
_FILL_ALT_ROW = PatternFill("solid", fgColor="F2F2F2")
_FONT_HEADER = Font(bold=True, color="FFFFFF", size=10)
_FONT_FLAGGED = Font(bold=True, color="C00000", size=10)
_ACCOUNTING_FMT = '#,##0.00_);(#,##0.00)'
_PCT_FMT = '0.00"%"'


# ---------------------------------------------------------------------------
# Core variance logic
# ---------------------------------------------------------------------------

def calculate_variances(
    current: pd.DataFrame,
    prior: pd.DataFrame,
    balance_col: str = "Balance",
    account_col: str = "Account_Number",
    name_col: str = "Account_Name",
    threshold_pct: float = 5.0,
) -> pd.DataFrame:
    """Merge two period DataFrames and compute account-level variances.

    Performs an outer join on *account_col* so accounts present in only one
    period are included with a zero balance for the missing period.

    Args:
        current:       DataFrame for the current period. Must contain *account_col*,
                       *name_col*, and *balance_col*.
        prior:         DataFrame for the prior period. Same column requirements.
        balance_col:   Name of the balance column in both DataFrames. Defaults
                       to ``"Balance"``.
        account_col:   Name of the account-identifier column. Defaults to
                       ``"Account_Number"``.
        name_col:      Name of the account description column. Defaults to
                       ``"Account_Name"``.
        threshold_pct: Absolute percentage-variance threshold above which an
                       account is flagged. Defaults to ``5.0`` (5 %).

    Returns:
        DataFrame with columns defined in ``_OUT_COLS``, sorted by
        ``abs(Variance_Amount)`` descending.

    Raises:
        ValueError: If a required column is missing from either DataFrame.
    """
    _validate_inputs(current, [account_col, name_col, balance_col], label="current")
    _validate_inputs(prior, [account_col, balance_col], label="prior")

    cur = current[[account_col, name_col, balance_col]].copy()
    cur.columns = ["Account_Number", "Account_Name", "Current_Balance"]

    pri = prior[[account_col, balance_col]].copy()
    pri.columns = ["Account_Number", "Prior_Balance"]

    return _build_report(cur, pri, threshold_pct)


def calculate_variances_single(
    df: pd.DataFrame,
    threshold_pct: float = 5.0,
) -> pd.DataFrame:
    """Compute variances from a single DataFrame with pre-split balance columns.

    Designed for DataFrames produced by ``data_parser.load_trial_balance``,
    which already contain both ``Current_Balance`` and ``Prior_Balance``.

    Args:
        df:            DataFrame with columns ``Account_Number``, ``Account_Name``,
                       ``Current_Balance``, and ``Prior_Balance``.
        threshold_pct: Absolute percentage-variance threshold. Defaults to ``5.0``.

    Returns:
        Variance report DataFrame sorted by ``abs(Variance_Amount)`` descending.

    Raises:
        ValueError: If any of the four required columns are missing.
    """
    required = ["Account_Number", "Account_Name", "Current_Balance", "Prior_Balance"]
    _validate_inputs(df, required, label="df")

    cur = df[["Account_Number", "Account_Name", "Current_Balance"]].copy()
    pri = df[["Account_Number", "Prior_Balance"]].copy()

    return _build_report(cur, pri, threshold_pct)


# ---------------------------------------------------------------------------
# Excel export
# ---------------------------------------------------------------------------

def export_to_excel(
    report: pd.DataFrame,
    output_path: str,
    sheet_name: str = "Variance Report",
) -> None:
    """Write the variance report to an Excel workbook with formatting.

    Formatting applied:
        - Dark-blue bold header row with white text.
        - Alternating light-grey fill on non-flagged data rows.
        - Red fill + bold red font on flagged rows, applied both directly and
          via an Excel FormulaRule so the highlighting survives in-cell edits.
        - Accounting number format on balance and variance columns.
        - Percentage format on the Variance_Pct column.
        - Column widths auto-fitted to content.
        - Top row frozen for scrolling.

    Args:
        report:      DataFrame produced by ``calculate_variances`` or
                     ``calculate_variances_single``.
        output_path: Destination ``.xlsx`` file path. Parent directories are
                     created automatically.
        sheet_name:  Name of the worksheet. Defaults to ``"Variance Report"``.

    Raises:
        ValueError: If *report* does not contain the expected columns.
        OSError:    If the file cannot be written (e.g. it is open in Excel).
    """
    _validate_inputs(report, _OUT_COLS, label="report")

    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    # Write raw data via pandas (handles the file creation cleanly)
    report.to_excel(path, sheet_name=sheet_name, index=False)
    log.info("Written %d rows to '%s'.", len(report), path)

    # Re-open with openpyxl for formatting
    wb = load_workbook(path)
    ws = wb[sheet_name]

    _format_header(ws)
    _format_data_rows(ws, report)
    _apply_conditional_formatting(ws, report)
    _set_column_widths(ws, report)
    ws.freeze_panes = "A2"

    wb.save(path)
    log.info("Formatting applied and workbook saved: '%s'.", path)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _validate_inputs(df: pd.DataFrame, required: list, label: str = "DataFrame") -> None:
    """Raise ValueError if any column in *required* is absent from *df*.

    Args:
        df:       DataFrame to check.
        required: List of column names that must be present.
        label:    Human-readable name used in the error message.

    Raises:
        ValueError: Lists all missing columns so the caller can fix them at once.
    """
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"'{label}' is missing required column(s): {missing}. "
            f"Found: {list(df.columns)}"
        )


def _build_report(
    cur: pd.DataFrame,
    pri: pd.DataFrame,
    threshold_pct: float,
) -> pd.DataFrame:
    """Merge, compute variance columns, flag, and sort.

    Args:
        cur:           DataFrame with Account_Number, Account_Name, Current_Balance.
        pri:           DataFrame with Account_Number, Prior_Balance.
        threshold_pct: Flag threshold (absolute percentage value).

    Returns:
        Finished variance report DataFrame.
    """
    merged = cur.merge(pri, on="Account_Number", how="outer")

    # Accounts missing from one period get a zero balance
    merged["Current_Balance"] = merged["Current_Balance"].fillna(0.0)
    merged["Prior_Balance"] = merged["Prior_Balance"].fillna(0.0)
    merged["Account_Name"] = merged["Account_Name"].fillna(merged["Account_Number"])

    merged["Variance_Amount"] = merged["Current_Balance"] - merged["Prior_Balance"]
    merged["Variance_Pct"] = merged.apply(_pct_change, axis=1)
    merged["Flagged"] = merged["Variance_Pct"].abs() > threshold_pct

    flagged_count = merged["Flagged"].sum()
    log.info(
        "Variance calculation complete: %d accounts, %d flagged (threshold: %.1f%%).",
        len(merged),
        flagged_count,
        threshold_pct,
    )
    if flagged_count:
        top = (
            merged[merged["Flagged"]]
            .nlargest(3, "Variance_Amount")["Account_Number"]
            .tolist()
        )
        log.debug("Top flagged accounts by variance amount: %s", top)

    return (
        merged[_OUT_COLS]
        .assign(Account_Number=merged["Account_Number"].astype(str))
        .sort_values("Variance_Amount", key=lambda s: s.abs(), ascending=False)
        .reset_index(drop=True)
    )


def _pct_change(row: pd.Series) -> float:
    """Compute percentage variance for a single account row.

    Returns 0.0 when both balances are zero, 100.0 when prior is zero
    but current is non-zero (new account), and the standard calculation
    otherwise.

    Args:
        row: Series with at least Current_Balance and Prior_Balance.

    Returns:
        Percentage change as a float (e.g. 12.5 means 12.5 %).
    """
    prior = row["Prior_Balance"]
    current = row["Current_Balance"]
    if prior == 0:
        return 0.0 if current == 0 else 100.0
    return (current - prior) / abs(prior) * 100.0


def _format_header(ws) -> None:
    """Apply dark-blue bold styling to the header row (row 1).

    Args:
        ws: openpyxl Worksheet object.
    """
    for cell in ws[1]:
        cell.fill = _FILL_HEADER
        cell.font = _FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _format_data_rows(ws, report: pd.DataFrame) -> None:
    """Apply per-row fills and number formats to data rows (rows 2+).

    Flagged rows receive red fill and red bold font. Non-flagged rows use
    alternating white / light-grey fills for readability.

    Args:
        ws:     openpyxl Worksheet object.
        report: The variance report DataFrame (used to read the Flagged column).
    """
    # Map output column names to their 1-based column indices in the worksheet
    col_index = {name: i + 1 for i, name in enumerate(_OUT_COLS)}

    balance_cols = {"Current_Balance", "Prior_Balance", "Variance_Amount"}
    pct_col = col_index["Variance_Pct"]

    for row_idx, (_, data_row) in enumerate(report.iterrows(), start=2):
        flagged = bool(data_row["Flagged"])
        fill = _FILL_FLAGGED if flagged else (_FILL_ALT_ROW if row_idx % 2 == 0 else None)
        font = _FONT_FLAGGED if flagged else Font(size=10)

        for col_name, col_i in col_index.items():
            cell = ws.cell(row=row_idx, column=col_i)
            if fill:
                cell.fill = fill
            cell.font = font

            if col_name in balance_cols:
                cell.number_format = _ACCOUNTING_FMT
            elif col_name == "Variance_Pct":
                cell.number_format = _PCT_FMT

    log.debug("Row-level formatting applied to %d data rows.", len(report))


def _apply_conditional_formatting(ws, report: pd.DataFrame) -> None:
    """Add an Excel FormulaRule so flagged rows stay highlighted after edits.

    The rule checks the ``Flagged`` column (column G) and colours any row
    where the cell equals TRUE. This works in addition to the static fill
    applied by ``_format_data_rows`` and ensures the workbook remains
    self-consistent if the user changes balance values after the fact.

    Args:
        ws:     openpyxl Worksheet object.
        report: Variance report DataFrame (used to determine the data range).
    """
    if report.empty:
        return

    last_row = len(report) + 1  # +1 for header
    last_col = get_column_letter(len(_OUT_COLS))
    data_range = f"A2:{last_col}{last_row}"

    # Column G is "Flagged" (7th column)
    flagged_col_letter = get_column_letter(_OUT_COLS.index("Flagged") + 1)

    rule = FormulaRule(
        formula=[f'${flagged_col_letter}2=TRUE'],
        fill=_FILL_FLAGGED,
        font=_FONT_FLAGGED,
    )
    ws.conditional_formatting.add(data_range, rule)
    log.debug("Conditional formatting rule applied to range %s.", data_range)


def _set_column_widths(ws, report: pd.DataFrame) -> None:
    """Set column widths based on the maximum content length in each column.

    Adds a small padding beyond the longest cell value. Caps width at 45
    characters to prevent excessively wide columns on long descriptions.

    Args:
        ws:     openpyxl Worksheet object.
        report: Variance report DataFrame.
    """
    for col_idx, col_name in enumerate(_OUT_COLS, start=1):
        header_len = len(col_name)
        try:
            max_data_len = report[col_name].astype(str).str.len().max()
        except Exception:
            max_data_len = 0

        width = min(max(header_len, int(max_data_len or 0)) + 4, 45)
        ws.column_dimensions[get_column_letter(col_idx)].width = width


# ---------------------------------------------------------------------------
# Standalone execution — quick sanity check
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    logging.basicConfig(
        level=logging.DEBUG,
        format="%(levelname)-8s %(name)s: %(message)s",
    )

    # Build a tiny in-memory example so the module can be tested without files.
    current_data = pd.DataFrame({
        "Account_Number": ["4000", "4100", "5000", "6000", "6100"],
        "Account_Name":   ["Product Sales", "Service Revenue", "Materials",
                           "Salaries", "Rent"],
        "Balance":        [178_000, 46_500, 72_000, 42_000, 8_000],
    })
    prior_data = pd.DataFrame({
        "Account_Number": ["4000", "4100", "5000", "6000", "6100"],
        "Account_Name":   ["Product Sales", "Service Revenue", "Materials",
                           "Salaries", "Rent"],
        "Balance":        [150_000, 45_000, 60_000, 35_000, 8_000],
    })

    threshold = float(sys.argv[1]) if len(sys.argv) > 1 else 5.0

    report = calculate_variances(current_data, prior_data, threshold_pct=threshold)
    print(report.to_string(index=False))

    out = sys.argv[2] if len(sys.argv) > 2 else "output/variance_report.xlsx"
    export_to_excel(report, out)
    print(f"\nExported to: {out}")
